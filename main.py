import os
import json
import threading
import requests
import re
import traceback
import shutil
import tempfile
from datetime import datetime
from flask import Flask
import telebot
from telebot import types

TOKEN = os.getenv("TOKEN")
bot = telebot.TeleBot(TOKEN)

app = Flask(__name__)

ACCESS_FILE = "access_data.json"
STATE_FILE = "user_states.json"

ADMIN_ID = 242294061

RENDER_API_KEY = os.getenv("RENDER_API_KEY")
RENDER_SERVICE_ID = os.getenv("RENDER_SERVICE_ID")
RENDER_PUBLIC_URL = os.getenv("RENDER_PUBLIC_URL", "https://tuo-servizio.onrender.com")

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4.1")

# GIF/video di benvenuto
WELCOME_MEDIA_PATH = os.getenv("WELCOME_MEDIA_PATH", "welcome.mp4")
WELCOME_MEDIA_URL = os.getenv("WELCOME_MEDIA_URL", "")
WELCOME_MEDIA_ENABLED = os.getenv("WELCOME_MEDIA_ENABLED", "true").lower() in {"1", "true", "yes", "si"}

PRESET_SCENARIOS = {
    "porto": {
        "label": "porto / terminal / crocieristi",
        "text": "veicolo ncc in area porto o terminal, possibile procacciamento clienti, possibile utenza indifferenziata, verificare prenotazione, foglio di servizio, rent, ruolo conducenti, kb, stazionamento su area pubblica"
    },
    "aeroporto": {
        "label": "aeroporto / arrivi / terminal",
        "text": "veicolo ncc in aeroporto o area arrivi, possibile attesa su area pubblica, possibile utenza indifferenziata, verificare prenotazione, foglio di servizio, rent, ruolo conducenti, kb"
    },
    "stazione": {
        "label": "stazione ferroviaria / terminal",
        "text": "veicolo ncc presso stazione ferroviaria o terminal, possibile procacciamento clienti, possibile utenza indifferenziata, verificare prenotazione, foglio di servizio, rent, ruolo conducenti, kb"
    },
    "hotel": {
        "label": "hotel / struttura ricettiva",
        "text": "servizio presso hotel o struttura ricettiva, verificare se navetta interna o servizio verso terzi, eventuale pagamento separato, prenotazione, foglio di servizio, rent, ruolo conducenti, kb"
    },
    "navetta": {
        "label": "navetta / parcheggio / shuttle",
        "text": "navetta o shuttle collegato a parcheggio o struttura, verificare se servizio accessorio o trasporto verso terzi, eventuale pagamento separato, prenotazione, foglio di servizio, rent, ruolo conducenti, kb"
    }
}



CONTROL_DOCS = [
    {"id": "patente", "label": "Patente"},
    {"id": "kb", "label": "KB / KA / CQC"},
    {"id": "autorizzazione", "label": "Licenza / Autorizzazione NCC"},
    {"id": "carta", "label": "Carta di circolazione / DU"},
    {"id": "assicurazione", "label": "Assicurazione"},
    {"id": "foglio", "label": "Foglio di servizio"},
]

CONTROL_ASSUME_NCC_DEFAULTS = {
    "vehicle_authorized": "si",
    "service_to_third": "si",
    "service_context": "a",
}

CONTROL_DOC_LABELS = {item["id"]: item["label"] for item in CONTROL_DOCS}

TARGHE_FILE_PATH = os.getenv("TARGHE_FILE_PATH", "prospetto_mezzi.xlsx")
TARGHE_SHEET_NAME = os.getenv("TARGHE_SHEET_NAME", "NCC")

ACCESS_DATA_FILE = os.getenv("ACCESS_DATA_FILE", "access_data.json")
USER_STATES_FILE = os.getenv("USER_STATES_FILE", "user_states.json")

# =========================
# DATI SANZIONI
# =========================

VIOLATIONS = {
    "085-02": {
        "title": "Noleggio con conducente di veicolo non adibito a tale uso - 1a violazione",
        "article": "CdS art. 85 c. 4",
        "pmr": "Non ammesso",
        "reduced_30": "Non ammesso",
        "over_60": "Non ammesso",
        "edictal": "da € 1.812,00 a € 7.249,00",
        "accessories": [
            "Sospensione patente da 4 a 12 mesi",
            "Sequestro del veicolo ai fini della confisca"
        ],
        "verbal_text": (
            "Adibiva a noleggio con conducente il veicolo sopra indicato benché non destinato a tale uso "
            "(ovvero, solo per autovetture e motocarrozzette, benché privo della prescritta autorizzazione; "
            "ovvero, per tutti i veicoli, benché l'autorizzazione fosse sospesa o revocata). "
            "Risultava infatti, in base alle indicazioni del documento di circolazione, che il veicolo doveva "
            "essere adibito a diverso uso. Nella circostanza veniva accertato il trasporto di persone. "
            "Il veicolo è sottoposto a sequestro ai fini della confisca come da separato verbale."
        ),
        "notes": [
            "Il verbale va inviato entro 10 giorni al Prefetto del luogo della violazione.",
            "Non è ammesso il pagamento in misura ridotta, essendo prevista la confisca del veicolo.",
            "Possibili violazioni concorrenti: CAP/CQC, art. 72, art. 79, 085-05.",
            "Non è ammesso il servizio NCC di autocarri per trasporto cose per conto terzi."
        ],
        "fields_to_fill": [
            "uso risultante dalla carta di circolazione / DU",
            "generalità di almeno un passeggero trasportato",
            "eventuale provvedimento di sospensione/revoca autorizzazione",
            "generalità del custode",
            "luogo di custodia del veicolo"
        ],
        "short_ready_text": (
            "Violazione accertata: noleggio con conducente con veicolo non adibito a tale uso. "
            "Norma: art. 85, comma 4, CdS. "
            "Sanzione edittale: da € 1.812,00 a € 7.249,00. "
            "Pagamento in misura ridotta non ammesso. "
            "Accessorie: sospensione patente da 4 a 12 mesi e sequestro del veicolo ai fini della confisca. "
            "Verbale da trasmettere entro 10 giorni al Prefetto del luogo della violazione."
        )
    },
    "085-04": {
        "title": "Noleggio con conducente di veicolo non adibito a tale uso - 2a violazione nel triennio",
        "article": "CdS art. 85 c. 4",
        "pmr": "Non ammesso",
        "reduced_30": "Non ammesso",
        "over_60": "Non ammesso",
        "edictal": "da € 1.812,00 a € 7.249,00",
        "accessories": [
            "Revoca patente",
            "Sequestro del veicolo per confisca"
        ],
        "verbal_text": (
            "Per la seconda volta, in un periodo di 3 anni, adibiva a noleggio con conducente il veicolo sopra indicato "
            "benché non destinato a tale uso (ovvero, solo per autovetture e motocarrozzette, benché fosse privo di "
            "autorizzazione ovvero, per tutti i veicoli, benché l'autorizzazione fosse sospesa o revocata). "
            "Risultava infatti, in base alle indicazioni del documento di circolazione, che il veicolo doveva essere "
            "adibito a diverso uso. Nella circostanza veniva accertato il trasporto di persone. "
            "Il veicolo è sottoposto a sequestro come da apposito verbale."
        ),
        "notes": [
            "Alla seconda violazione nel triennio si applica la revoca patente.",
            "L'organo accertatore comunica entro 5 giorni al Prefetto i presupposti della revoca."
        ],
        "fields_to_fill": [
            "uso risultante dalla carta di circolazione / DU",
            "generalità di almeno un passeggero trasportato",
            "generalità del custode",
            "luogo di custodia del veicolo",
            "precedente violazione utile nel triennio"
        ],
        "short_ready_text": (
            "Violazione accertata: noleggio con conducente con veicolo non adibito a tale uso, seconda violazione nel triennio. "
            "Norma: art. 85, comma 4, CdS. "
            "Sanzione edittale: da € 1.812,00 a € 7.249,00. "
            "Pagamento in misura ridotta non ammesso. "
            "Accessorie: revoca patente e sequestro del veicolo per confisca. "
            "Comunicazione entro 5 giorni al Prefetto per i presupposti della revoca."
        )
    },
    "085-05": {
        "title": "NCC in violazione degli artt. 3 e 11 L. 21/1992 - 1a violazione nel quinquennio",
        "article": "CdS art. 85 c. 4-bis + L. 21/1992",
        "pmr": "€ 178,00",
        "reduced_30": "€ 124,60",
        "over_60": "€ 336,00",
        "edictal": "da € 178,00 a € 672,00",
        "accessories": [
            "Sospensione documento di circolazione per 1 mese"
        ],
        "verbal_text": (
            "Utilizzava l'autovettura (o il motoveicolo) sopra indicata, adibita a servizio di noleggio con conducente, "
            "munito della relativa autorizzazione, senza ottemperare alle disposizioni dell'art. 3 della L. 21/1992 "
            "ovvero alle disposizioni di cui all'art. 11 della L. 21/1992. "
            "Il documento di circolazione è ritirato e trasmesso all'UMC competente. "
            "Il conducente è autorizzato a condurre il veicolo per la via più breve fino al luogo indicato, "
            "con l'avvertenza che lo stesso sarà sottoposto a fermo amministrativo per tutta la durata della sospensione."
        ),
        "notes": [
            "Se il conducente non è titolare dell’autorizzazione, la violazione va notificata al titolare.",
            "Non contestare come autonoma violazione il mero mancato rientro in rimessa dopo ogni servizio.",
            "Verificare con attenzione prenotazione, stazionamento, foglio di servizio, sede/rimessa."
        ],
        "fields_to_fill": [
            "specificare se la violazione riguarda art. 3 o art. 11 L. 21/1992",
            "descrizione concreta del fatto accertato",
            "UMC competente",
            "luogo verso cui autorizzare la marcia per la via più breve",
            "titolare dell'autorizzazione se diverso dal conducente"
        ],
        "short_ready_text": (
            "Violazione accertata: utilizzo del veicolo NCC in violazione degli artt. 3 o 11 della L. 21/1992. "
            "Norma: art. 85, comma 4-bis, CdS. "
            "PMR € 178,00; riduzione 30% € 124,60; oltre 60 giorni € 336,00; "
            "edittale da € 178,00 a € 672,00. "
            "Accessoria: sospensione del documento di circolazione per 1 mese."
        )
    },
    "085-06": {
        "title": "NCC in violazione degli artt. 3 e 11 L. 21/1992 - 2a violazione nel quinquennio",
        "article": "CdS art. 85 c. 4-bis + L. 21/1992",
        "pmr": "€ 264,00",
        "reduced_30": "€ 184,80",
        "over_60": "€ 505,00",
        "edictal": "da € 264,00 a € 1.010,00",
        "accessories": [
            "Sospensione documento di circolazione da 1 a 2 mesi"
        ],
        "verbal_text": (
            "Utilizzava l'autovettura (o il motoveicolo) sopra indicata, adibita a servizio di noleggio con conducente, "
            "munito della relativa autorizzazione, senza ottemperare alle disposizioni dell'art. 3 della L. 21/1992 "
            "ovvero alle disposizioni di cui all'art. 11 della L. 21/1992. "
            "Il documento di circolazione è ritirato e trasmesso all'UMC competente."
        ),
        "notes": [
            "Seconda violazione nel quinquennio.",
            "Se il conducente non è titolare, risponde il titolare dell’autorizzazione."
        ],
        "fields_to_fill": [
            "specificare se la violazione riguarda art. 3 o art. 11 L. 21/1992",
            "descrizione concreta del fatto accertato",
            "UMC competente",
            "precedente violazione utile nel quinquennio"
        ],
        "short_ready_text": (
            "Violazione accertata: utilizzo del veicolo NCC in violazione degli artt. 3 o 11 della L. 21/1992, "
            "seconda violazione nel quinquennio. "
            "Norma: art. 85, comma 4-bis, CdS. "
            "PMR € 264,00; riduzione 30% € 184,80; oltre 60 giorni € 505,00; "
            "edittale da € 264,00 a € 1.010,00. "
            "Accessoria: sospensione del documento di circolazione da 1 a 2 mesi."
        )
    },
    "085-07": {
        "title": "NCC in violazione degli artt. 3 e 11 L. 21/1992 - 3a violazione nel quinquennio",
        "article": "CdS art. 85 c. 4-bis + L. 21/1992",
        "pmr": "€ 356,00",
        "reduced_30": "€ 249,20",
        "over_60": "€ 672,00",
        "edictal": "da € 356,00 a € 1.344,00",
        "accessories": [
            "Sospensione documento di circolazione da 2 a 4 mesi"
        ],
        "verbal_text": (
            "Utilizzava l'autovettura (o il motoveicolo) sopra indicata, adibita a servizio di noleggio con conducente, "
            "munito della relativa autorizzazione, senza ottemperare alle disposizioni dell'art. 3 della L. 21/1992 "
            "ovvero alle disposizioni di cui all'art. 11 della L. 21/1992."
        ),
        "notes": [
            "Terza violazione nel quinquennio."
        ],
        "fields_to_fill": [
            "specificare se la violazione riguarda art. 3 o art. 11 L. 21/1992",
            "descrizione concreta del fatto accertato",
            "precedenti violazioni utili nel quinquennio"
        ],
        "short_ready_text": (
            "Violazione accertata: utilizzo del veicolo NCC in violazione degli artt. 3 o 11 della L. 21/1992, "
            "terza violazione nel quinquennio. "
            "Norma: art. 85, comma 4-bis, CdS. "
            "PMR € 356,00; riduzione 30% € 249,20; oltre 60 giorni € 672,00; "
            "edittale da € 356,00 a € 1.344,00. "
            "Accessoria: sospensione del documento di circolazione da 2 a 4 mesi."
        )
    },
    "085-08": {
        "title": "NCC in violazione degli artt. 3 e 11 L. 21/1992 - 4a o successiva nel quinquennio",
        "article": "CdS art. 85 c. 4-bis + L. 21/1992",
        "pmr": "€ 528,00",
        "reduced_30": "€ 369,60",
        "over_60": "€ 1.010,00",
        "edictal": "da € 528,00 a € 2.020,00",
        "accessories": [
            "Sospensione documento di circolazione da 4 a 8 mesi"
        ],
        "verbal_text": (
            "Utilizzava l'autovettura (o il motoveicolo) sopra indicata, adibita a servizio di noleggio con conducente, "
            "munito della relativa autorizzazione, senza ottemperare alle disposizioni dell'art. 3 della L. 21/1992 "
            "ovvero alle disposizioni di cui all'art. 11 della L. 21/1992."
        ),
        "notes": [
            "Quarta o successiva violazione nel quinquennio."
        ],
        "fields_to_fill": [
            "specificare se la violazione riguarda art. 3 o art. 11 L. 21/1992",
            "descrizione concreta del fatto accertato",
            "progressione completa delle violazioni nel quinquennio"
        ],
        "short_ready_text": (
            "Violazione accertata: utilizzo del veicolo NCC in violazione degli artt. 3 o 11 della L. 21/1992, "
            "quarta o successiva nel quinquennio. "
            "Norma: art. 85, comma 4-bis, CdS. "
            "PMR € 528,00; riduzione 30% € 369,60; oltre 60 giorni € 1.010,00; "
            "edittale da € 528,00 a € 2.020,00. "
            "Accessoria: sospensione del documento di circolazione da 4 a 8 mesi."
        )
    },
    "085-09": {
        "title": "Circolazione con NCC violando altre prescrizioni dell’autorizzazione",
        "article": "CdS art. 85 c. 4-ter",
        "pmr": "€ 86,00",
        "reduced_30": "€ 60,20",
        "over_60": "€ 169,00",
        "edictal": "da € 86,00 a € 338,00",
        "accessories": [
            "Non previste"
        ],
        "verbal_text": (
            "Utilizzava il veicolo sopra indicato, adibito a servizio di noleggio con conducente, "
            "senza ottemperare alle norme in vigore ovvero alle condizioni di cui all'autorizzazione. "
            "Veniva accertato che ..."
        ),
        "notes": [
            "Rientrano qui le prescrizioni diverse dagli artt. 3 e 11 L. 21/1992.",
            "Verificare regolamenti comunali/locali, ZTL, modalità di servizio, ruolo CCIAA."
        ],
        "fields_to_fill": [
            "specifica prescrizione autorizzativa violata",
            "descrizione concreta del fatto accertato",
            "eventuale regolamento locale applicabile"
        ],
        "short_ready_text": (
            "Violazione accertata: circolazione con NCC in violazione di altre prescrizioni dell'autorizzazione. "
            "Norma: art. 85, comma 4-ter, CdS. "
            "PMR € 86,00; riduzione 30% € 60,20; oltre 60 giorni € 169,00; "
            "edittale da € 86,00 a € 338,00. "
            "Sanzioni accessorie non previste."
        )
    },
    "116-06": {
        "title": "Guida senza CAP o CQC",
        "article": "CdS art. 116 c. 16 e 18",
        "pmr": "€ 408,00",
        "reduced_30": "€ 285,60",
        "over_60": "€ 817,00",
        "edictal": "da € 408,00 a € 1.634,00",
        "accessories": [
            "Fermo veicolo per 60 giorni"
        ],
        "verbal_text": (
            "Circolava alla guida del predetto veicolo adibito a servizio di noleggio con conducente "
            "munito di patente ma non del prescritto certificato di abilitazione professionale / titolo "
            "professionale richiesto per il servizio svolto."
        ),
        "notes": [
            "Da valutare il concorso con le violazioni ex art. 85 CdS.",
            "Verificare se per il servizio concreto fosse richiesto KB / KA / CQC."
        ],
        "fields_to_fill": [
            "titolo professionale mancante (KB / KA / CQC)",
            "tipo di veicolo e servizio svolto",
            "eventuale affidante del veicolo"
        ],
        "short_ready_text": (
            "Violazione concorrente: guida del veicolo adibito al servizio senza il prescritto titolo professionale. "
            "Norma: art. 116, commi 16 e 18, CdS. "
            "PMR € 408,00; riduzione 30% € 285,60; oltre 60 giorni € 817,00; "
            "edittale da € 408,00 a € 1.634,00. "
            "Accessoria: fermo del veicolo per 60 giorni."
        )
    }
}

VIOLATIONS["180-01"] = {
    "title": "Mancata esibizione del foglio di servizio / documento di controllo",
    "article": "CdS art. 180 (valutazione separata)",
    "pmr": "Verificare prontuario",
    "reduced_30": "Verificare prontuario",
    "over_60": "Verificare prontuario",
    "edictal": "Verificare fattispecie concreta",
    "accessories": ["Verificare"],
    "verbal_text": "Il conducente non esibiva nell'immediatezza il foglio di servizio o il codice identificativo del servizio richiesto in controllo.",
    "notes": ["Da usare come richiamo operativo separato e solo se la fattispecie concreta è documentale/non sostanziale."],
    "fields_to_fill": ["documento richiesto", "modalità della richiesta", "esito del controllo"],
    "short_ready_text": "Valutare autonoma contestazione documentale per mancata esibizione immediata del foglio di servizio/codice identificativo del servizio."
}


VIOLATIONS["180-DOC"] = {
    "title": "Mancata esibizione di documento obbligatorio a bordo",
    "article": "CdS art. 180 (da definire in base al documento)",
    "pmr": "Verificare prontuario",
    "reduced_30": "Verificare prontuario",
    "over_60": "Verificare prontuario",
    "edictal": "Verificare fattispecie concreta",
    "accessories": ["Verificare"],
    "verbal_text": "Il conducente non esibiva all'atto del controllo uno o più documenti obbligatori per la circolazione o per il servizio.",
    "notes": [
        "Usare come richiamo operativo per patente, autorizzazione NCC, carta di circolazione o assicurazione non esibite.",
        "La voce specifica di prontuario va individuata in base al documento mancante all'atto del controllo."
    ],
    "fields_to_fill": ["documento richiesto", "modalità della richiesta", "ufficio di presentazione se previsto"],
    "short_ready_text": "Valutare autonoma contestazione documentale ex art. 180 CdS per i documenti non esibiti all'atto del controllo."
}

VIOLATIONS["180-01DOC"] = {
    "title": "Mancanza momentanea di documenti di guida/circolazione",
    "article": "CdS art. 180 c. 1 e c. 7",
    "pmr": "€ 42,00",
    "reduced_30": "€ 29,40",
    "over_60": "€ 86,50",
    "edictal": "da € 42,00 a € 173,00",
    "accessories": ["Nessuna"],
    "verbal_text": "Circolava alla guida del veicolo sopra indicato senza portare con sé, pur dichiarando di esserne in possesso, la patente di guida valida per la categoria del veicolo condotto e/o il documento di circolazione. L'interessato è invitato a presentarsi entro 30 giorni presso un ufficio di polizia per esibire il documento mancante. L'inosservanza a quanto intimato comporterà la sanzione di cui all'art. 180 c. 8 CdS.",
    "notes": [
        "Usare questa voce per patente o carta di circolazione non al seguito quando il documento esiste.",
        "Se alla verifica il documento risulta inesistente o revocato, applicare la norma specifica e non l'art. 180."
    ],
    "fields_to_fill": ["documento mancante", "situazione di fatto da precisare", "ufficio di polizia", "termine per presentazione"],
    "short_ready_text": "Art. 180 c.1 e c.7: PMR € 42,00; riduzione 30% € 29,40; oltre 60 giorni € 86,50; edittale da € 42,00 a € 173,00. Invito a presentare il documento mancante entro il termine fissato."
}

VIOLATIONS["180-03"] = {
    "title": "Mancanza momentanea del certificato assicurativo",
    "article": "CdS art. 180 c. 1 e c. 7",
    "pmr": "€ 42,00",
    "reduced_30": "€ 29,40",
    "over_60": "€ 86,50",
    "edictal": "da € 42,00 a € 173,00",
    "accessories": ["Nessuna"],
    "verbal_text": "Conducente del veicolo sopra indicato non aveva con sé il certificato di assicurazione obbligatoria. Avendo l'utente dichiarato l'esistenza dell'assicurazione con la compagnia/agenzia da indicare, l'interessato è invitato a presentarsi entro 30 giorni presso un ufficio di polizia per esibire il documento mancante. L'inosservanza a quanto intimato comporterà la sanzione di cui all'art. 180 c. 8 CdS.",
    "notes": [
        "Usare solo se la copertura esiste ed è solo non esibita.",
        "Se la copertura manca, applicare l'art. 193."
    ],
    "fields_to_fill": ["compagnia assicurativa", "agenzia", "motivo dell'impossibilità di accertamento immediato", "ufficio di polizia"],
    "short_ready_text": "Art. 180 c.1 e c.7 per certificato assicurativo non al seguito: PMR € 42,00; riduzione 30% € 29,40; oltre 60 giorni € 86,50. Invito a presentare il documento mancante."
}

VIOLATIONS["180-06"] = {
    "title": "Mancanza momentanea di autorizzazioni, licenze o altri documenti",
    "article": "CdS art. 180 c. 3 e c. 7",
    "pmr": "€ 42,00",
    "reduced_30": "€ 29,40",
    "over_60": "€ 86,50",
    "edictal": "da € 42,00 a € 173,00",
    "accessories": ["Nessuna"],
    "verbal_text": "Conducente del veicolo sopra indicato, adibito a servizio NCC, non portava con sé la specifica autorizzazione o licenza prescritta. L'interessato è invitato a presentarsi entro 30 giorni presso un ufficio di polizia per esibire il documento mancante. L'inosservanza a quanto intimato comporterà la sanzione di cui all'art. 180 c. 8 CdS.",
    "notes": [
        "Per i veicoli adibiti agli usi previsti dall'art. 82 il conducente deve avere con sé la specifica autorizzazione/licenza.",
        "Se il titolo risulta inesistente, sospeso o revocato, applicare la norma specifica del settore e non l'art. 180."
    ],
    "fields_to_fill": ["tipo di servizio", "licenza/autorizzazione mancante", "ufficio di polizia", "termine per presentazione"],
    "short_ready_text": "Art. 180 c.3 e c.7 per licenza/autorizzazione NCC non al seguito: PMR € 42,00; riduzione 30% € 29,40; oltre 60 giorni € 86,50. Invito a esibire il documento mancante."
}

VIOLATIONS["180-09"] = {
    "title": "Mancanza di CAP, CQC, CFP o certificato di idoneità",
    "article": "CdS art. 180 c. 5 e c. 7",
    "pmr": "€ 42,00",
    "reduced_30": "€ 29,40",
    "over_60": "€ 86,50",
    "edictal": "da € 42,00 a € 173,00",
    "accessories": ["Nessuna"],
    "verbal_text": "Conducente del veicolo sopra indicato, per il quale era prescritto il possesso di CAP o CQC, non portava con sé tale documento. L'interessato è invitato a presentarsi entro 30 giorni presso un ufficio di polizia per esibire il documento mancante. L'inosservanza a quanto intimato comporterà la sanzione di cui all'art. 180 c. 8 CdS.",
    "notes": [
        "Usare se il titolo professionale esiste ma non è stato esibito.",
        "Se il titolo manca o non è idoneo, applicare l'art. 116 c. 16 e 18."
    ],
    "fields_to_fill": ["titolo mancante", "ufficio di polizia", "termine per presentazione"],
    "short_ready_text": "Art. 180 c.5 e c.7 per CAP/KB/CQC non al seguito: PMR € 42,00; riduzione 30% € 29,40; oltre 60 giorni € 86,50. Invito a presentare il titolo mancante."
}

NCC_DB = {
    "norme": {
        "L21_3_11": {
            "norma": "L. 21/1992 artt. 3 e 11",
            "tema": "Prescrizioni autorizzazione/licenza NCC",
            "uso_operativo": "Verificare se la violazione ricade nelle prescrizioni tipiche del titolo NCC e nel corretto esercizio del servizio."
        },
        "CDS_85_4": {
            "norma": "CdS art. 85 c. 4",
            "tema": "Veicolo non adibito/autorizzato a NCC",
            "uso_operativo": "Ramo sanzionatorio quando il servizio è svolto con veicolo non autorizzato/adibito a tale uso."
        },
        "CDS_85_4BIS": {
            "norma": "CdS art. 85 c. 4-bis",
            "tema": "Violazione artt. 3 o 11 L. 21/1992",
            "uso_operativo": "Ramo per violazioni collegate alle prescrizioni della L. 21/1992, con progressione nel quinquennio."
        },
        "CDS_85_4TER": {
            "norma": "CdS art. 85 c. 4-ter",
            "tema": "Altre prescrizioni autorizzazione NCC",
            "uso_operativo": "Ramo per violazioni di prescrizioni diverse da artt. 3 e 11 L. 21/1992."
        },
        "CDS_116_14": {
            "norma": "CdS art. 116 c. 14",
            "tema": "Incauto affidamento",
            "uso_operativo": "Quando il soggetto che ha disponibilità del veicolo lo affida a persona priva dei titoli richiesti."
        },
        "CDS_116_15_17": {
            "norma": "CdS art. 116 c. 15 e 17",
            "tema": "Guida senza patente o patente non idonea",
            "uso_operativo": "Prima violazione amministrativa, recidiva biennale o reiterazione; può concorrere con il ramo NCC."
        },
        "CDS_116_15BIS": {
            "norma": "CdS art. 116 c. 15-bis",
            "tema": "Patente diversa in casi tipizzati",
            "uso_operativo": "Ipotesi specifica distinta dalla guida senza patente vera e propria."
        },
        "CDS_116_16_18": {
            "norma": "CdS art. 116 c. 16 e 18",
            "tema": "Guida senza CAP/KB/KA/CQC",
            "uso_operativo": "Per servizio pubblico/NCC senza titolo professionale richiesto."
        },
        "CDS_126_11": {
            "norma": "CdS art. 126 c. 11",
            "tema": "CAP/CQC scaduti",
            "uso_operativo": "Quando il titolo professionale esiste ma non è più valido."
        },
        "CDS_180": {
            "norma": "CdS art. 180",
            "tema": "Mancanza materiale del documento",
            "uso_operativo": "Solo quando il titolo esiste ma non è esibito nell’immediato."
        }
    },

    "voci": {
        "085-02": {
            "norma": "CdS art. 85 c. 4",
            "titolo": "Veicolo non adibito a tale uso – 1ª violazione",
            "quando_usarla": "Servizio NCC svolto con veicolo non regolarmente adibito/autorizzato a NCC.",
            "cluster": "mezzo_non_ncc"
        },
        "085-04": {
            "norma": "CdS art. 85 c. 4",
            "titolo": "Veicolo non adibito – 2ª violazione nel triennio",
            "quando_usarla": "Quando ricorre seconda violazione nel triennio sul ramo del veicolo non autorizzato.",
            "cluster": "mezzo_non_ncc"
        },
        "085-05": {
            "norma": "CdS art. 85 c. 4-bis + L. 21/1992",
            "titolo": "Violazione artt. 3 e 11 – 1ª nel quinquennio",
            "quando_usarla": "Quando il veicolo è NCC e la violazione riguarda artt. 3 o 11.",
            "cluster": "mezzo_ncc_art3_11"
        },
        "085-06": {
            "norma": "CdS art. 85 c. 4-bis + L. 21/1992",
            "titolo": "Violazione artt. 3 e 11 – 2ª nel quinquennio",
            "quando_usarla": "Come sopra, con progressione sanzionatoria.",
            "cluster": "mezzo_ncc_art3_11"
        },
        "085-07": {
            "norma": "CdS art. 85 c. 4-bis + L. 21/1992",
            "titolo": "Violazione artt. 3 e 11 – 3ª nel quinquennio",
            "quando_usarla": "Come sopra.",
            "cluster": "mezzo_ncc_art3_11"
        },
        "085-08": {
            "norma": "CdS art. 85 c. 4-bis + L. 21/1992",
            "titolo": "Violazione artt. 3 e 11 – 4ª o successiva",
            "quando_usarla": "Come sopra.",
            "cluster": "mezzo_ncc_art3_11"
        },
        "085-09": {
            "norma": "CdS art. 85 c. 4-ter",
            "titolo": "Altre prescrizioni dell’autorizzazione NCC",
            "quando_usarla": "Quando la violazione non ricade negli artt. 3 e 11 ma in altre prescrizioni dell’autorizzazione.",
            "cluster": "mezzo_ncc_altre_prescrizioni"
        },
        "116-01": {
            "norma": "CdS art. 116 c. 14",
            "titolo": "Incauto affidamento",
            "quando_usarla": "Veicolo affidato a soggetto privo dei titoli richiesti.",
            "cluster": "conducente"
        },
        "116-02": {
            "norma": "CdS art. 116 c. 15 e 17",
            "titolo": "Guida senza patente / prima violazione amministrativa",
            "quando_usarla": "Mai conseguita, revocata, non rinnovata per mancanza requisiti o categoria non rientrante nel 15-bis.",
            "cluster": "conducente"
        },
        "116-03": {
            "norma": "CdS art. 116 c. 15 e 17",
            "titolo": "Recidiva biennale nella violazione amministrativa",
            "quando_usarla": "Seconda violazione amministrativa nel biennio, senza presupposti della reiterazione penale.",
            "cluster": "conducente"
        },
        "116-04": {
            "norma": "CdS art. 116 c. 15 e 17",
            "titolo": "Reiterazione nella guida senza patente",
            "quando_usarla": "Quando ricorrono i presupposti di illecito penale richiamati dal prontuario.",
            "cluster": "conducente"
        },
        "116-05": {
            "norma": "CdS art. 116 c. 15-bis",
            "titolo": "Patente diversa ma rientrante nei casi tipizzati",
            "quando_usarla": "Ipotesi tassative di patente di categoria diversa.",
            "cluster": "conducente"
        },
        "116-06": {
            "norma": "CdS art. 116 c. 16 e 18",
            "titolo": "Guida senza CAP o CQC",
            "quando_usarla": "Servizio NCC senza KB/KA/CQC quando richiesti.",
            "cluster": "conducente"
        }
    },

    "casi_tipici": {
        "A_abusivo_totale_mezzo_proprio": {
            "descrizione": "Un privato usa il proprio veicolo per trasporto verso terzi assimilabile a NCC.",
            "possibili_esiti": ["085-02", "085-04", "116-02", "116-03", "116-04", "116-05", "116-06"],
            "note": [
                "Se manca il corretto inquadramento NCC del veicolo: valutare il ramo art. 85 c. 4.",
                "Se manca patente idonea: valutare art. 116 c. 15 / 15-bis secondo il caso.",
                "Se manca il titolo professionale richiesto: valutare 116-06.",
                "Possibile concorso di più violazioni."
            ]
        },
        "B_titolare_ncc_procaccia_clienti": {
            "descrizione": "Il mezzo è NCC ma il servizio va verificato rispetto alle prescrizioni della L. 21/1992 e dell’autorizzazione.",
            "possibili_esiti": ["085-05", "085-06", "085-07", "085-08", "085-09", "116-06"],
            "note": [
                "Se la violazione riguarda artt. 3 o 11 L. 21/1992: ramo 085-05 / 085-06 / 085-07 / 085-08.",
                "Se riguarda altre prescrizioni dell’autorizzazione: 085-09.",
                "Se il conducente non ha titolo professionale valido: può concorrere 116-06."
            ]
        },
        "C_impresa_agenzia_mezzo_aziendale": {
            "descrizione": "Impresa/agenzia che trasporta clienti con dipendente e mezzo aziendale.",
            "possibili_esiti": ["085-02", "085-04", "085-05", "085-09", "116-06"],
            "note": [
                "Prima domanda pratica: trasporto verso terzi / assimilabile a NCC oppure attività accessoria dell’impresa?",
                "Se il servizio è organizzato come trasporto a terzi, il rischio è la qualificazione come servizio NCC non autorizzato.",
                "Se il trasporto è realmente accessorio al servizio principale dell’impresa e riservato ai propri clienti, la qualificazione va verificata con attenzione."
            ]
        },
        "D_navetta_parcheggio_hotel": {
            "descrizione": "Navetta interna / parcheggio / hotel / struttura.",
            "possibili_esiti": ["085-02", "085-04"],
            "note": [
                "Va distinta la navetta interna/collegamento funzionale dal trasporto verso terzi autonomamente venduto.",
                "Se il trasporto è compreso nel servizio principale e dedicato ai soli clienti, il caso va valutato come attività accessoria e non automaticamente come NCC.",
                "Se invece il trasporto è offerto come servizio a terzi autonomo, aumenta il rischio di riconduzione al ramo NCC abusivo.",
                "Verificare copertura documentale e contrattuale."
            ]
        },
        "E_veicolo_ncc_conducente_non_legittimato": {
            "descrizione": "Il veicolo risulta NCC, ma il conducente non ha patente idonea o titolo professionale richiesto.",
            "possibili_esiti": ["116-01", "116-02", "116-03", "116-04", "116-05", "116-06"],
            "note": [
                "Il nucleo sanzionatorio principale è spesso sul ramo 116.",
                "Può aggiungersi l’incauto affidamento 116-01 a carico di chi ha affidato il veicolo."
            ]
        },
        "F_veicolo_non_ncc_ma_conducente_titolato": {
            "descrizione": "Il conducente ha titolo ma usa mezzo non autorizzato/adibito a NCC.",
            "possibili_esiti": ["085-02", "085-04", "116-02", "116-03", "116-04", "116-05", "116-06"],
            "note": [
                "Il fatto che il conducente abbia titolo non sana l’uso di un mezzo non autorizzato/adibito a NCC.",
                "Valutare prioritariamente il ramo art. 85 c. 4.",
                "Se inoltre mancano patente idonea o requisiti documentali, il ramo 116 può concorrere."
            ]
        }
    },

    "checklist_operativa": [
        "Capire se il trasporto è verso terzi / assimilabile a NCC oppure attività accessoria dell’impresa.",
        "Identificare il soggetto che svolge il trasporto: privato, titolare NCC, impresa/agenzia, gestore navetta/parcheggio.",
        "Verificare come viene acquisita la clientela: prenotazione preventiva, procacciamento diretto, servizio per clienti propri, navetta interna.",
        "Verificare se il servizio è riservato ai clienti della struttura o è offerto come trasporto a terzi.",
        "Verificare se il veicolo è regolarmente adibito/autorizzato a NCC.",
        "Verificare patente idonea del conducente.",
        "Verificare titolo professionale richiesto (KB/KA/CQC) o eventuale scadenza.",
        "Se il veicolo è NCC, distinguere: violazione artt. 3 o 11 L. 21/1992 / altre prescrizioni / progressione nel quinquennio.",
        "Valutare eventuale concorso con incauto affidamento nei confronti dell’avente disponibilità del veicolo."
    ],

    "matrice_rapida": {
        "Mezzo non autorizzato NCC": ["085-02", "085-04"],
        "Mezzo NCC + violazione artt. 3 o 11 L. 21/1992": ["085-05", "085-06", "085-07", "085-08"],
        "Mezzo NCC + altre prescrizioni autorizzazione": ["085-09"],
        "Conducente senza patente idonea": ["116-02", "116-03", "116-04", "116-05"],
        "Conducente senza KB/KA/CQC": ["116-06"],
        "Veicolo affidato a soggetto privo dei titoli richiesti": ["116-01"],
        "Titolo professionale esistente ma scaduto": ["CDS_126_11"],
        "Documento esistente ma non esibito": ["CDS_180"]
    },

    "documenti_controllo": [
        "Carta di circolazione / Documento Unico di circolazione e proprietà (DU)",
        "Autorizzazione NCC / titolo abilitativo",
        "Patente del conducente",
        "CAP (KB) / KA / CQC se richiesti",
        "Iscrizione a ruolo prov. dei conducenti (CCIIAA)/ posizione del conducente",
        "Prenotazione documentabile",
        "Foglio di servizio compilato prima di ogni corsa",
        "Eventuale documentazione contrattuale / commerciale / fiscale",
        "Autorizzazione Comunale / accesso porto / ZTL"
    ]
}

ARTICOLI_DB = {
    "art85": {
        "titolo": "CdS art. 85",
        "testo": (
            "Art. 85 CdS – Servizio di noleggio con conducente.\n\n"
            "Uso operativo nel bot:\n"
            "- comma 4: veicolo non adibito / non autorizzato a NCC\n"
            "- comma 4-bis: violazioni degli artt. 3 e 11 della L. 21/1992\n"
            "- comma 4-ter: altre prescrizioni dell’autorizzazione NCC\n\n"
            "Per il dettaglio sanzionatorio il bot usa le voci:\n"
            "085-02, 085-04, 085-05, 085-06, 085-07, 085-08, 085-09."
        ),
        "link": "https://www.normattiva.it/"
    },
    "art116": {
        "titolo": "CdS art. 116",
        "testo": (
            "Art. 116 CdS – Requisiti per la guida dei veicoli.\n\n"
            "Uso operativo nel bot:\n"
            "- comma 14: incauto affidamento\n"
            "- commi 15 e 17: guida senza patente / recidiva / reiterazione\n"
            "- comma 15-bis: patente diversa in casi tipizzati\n"
            "- commi 16 e 18: guida senza CAP / KB / KA / CQC\n\n"
            "Per il dettaglio operativo il bot usa le voci:\n"
            "116-01, 116-02, 116-03, 116-04, 116-05, 116-06."
        ),
        "link": "https://www.normattiva.it/"
    },
    "art3l21": {
        "titolo": "L. 21/1992 art. 3",
        "testo": (
            "Legge 21/1992 – art. 3.\n\n"
            "Uso operativo nel bot:\n"
            "- la richiesta del servizio deve essere avanzata presso la sede o la rimessa, "
            "anche mediante strumenti tecnologici;\n"
            "- lo stazionamento dei mezzi deve avvenire all’interno delle rimesse;\n"
            "- sede operativa e almeno una rimessa devono essere nel territorio consentito.\n\n"
            "Se il mezzo è regolarmente NCC ma viola queste prescrizioni, il bot orienta verso il ramo 085-05 / 085-06 / 085-07 / 085-08."
        ),
        "link": "https://www.normattiva.it/"
    },
    "art11l21": {
        "titolo": "L. 21/1992 art. 11",
        "testo": (
            "Legge 21/1992 – art. 11.\n\n"
            "Uso operativo nel bot:\n"
            "- divieto di sosta in posteggio di stazionamento su suolo pubblico nei comuni dove è esercitato il servizio taxi;\n"
            "- nei casi consentiti, verifica dell’area di sosta autorizzata;\n"
            "- controllo su prenotazione / foglio di servizio / modalità operative del servizio.\n\n"
            "Se il mezzo è NCC e la violazione ricade su queste prescrizioni, il bot orienta verso il ramo 085-05 / 085-06 / 085-07 / 085-08."
        ),
        "link": "https://www.normattiva.it/"
 
    },
    "art180": {
        "titolo": "CdS art. 180",
        "testo": (
            "Art. 180 CdS – Possesso ed esibizione dei documenti di circolazione e di guida.\n\n"
            "Uso operativo nel bot:\n"
            "- si applica quando il documento esiste ma non viene esibito all'atto del controllo;\n"
            "- va distinto dalle violazioni sostanziali, cioe quando il titolo manca, e revocato, sospeso o non idoneo;\n"
            "- nel bot richiama soprattutto le ipotesi documentali per patente, autorizzazione NCC, foglio di servizio, carta di circolazione e assicurazione."
        ),
        "link": "https://www.normattiva.it/"
    },
    "art126": {
        "titolo": "CdS art. 126",
        "testo": (
            "Art. 126 CdS – Durata e conferma della validita della patente di guida.\n\n"
            "Uso operativo nel bot:\n"
            "- rileva quando patente, CAP, KB o CQC risultano scaduti;\n"
            "- va distinto dalla mancanza del titolo e dalla mera mancata esibizione;\n"
            "- nel bot e usato come richiamo operativo per i titoli scaduti."
        ),
        "link": "https://www.normattiva.it/"
   }
}

# =========================
# ACCESSO / AUTORIZZAZIONI
# =========================

ACCESS_DATA_BACKUP_FILE = os.getenv("ACCESS_DATA_BACKUP_FILE", f"{ACCESS_DATA_FILE}.bak")
BOOTSTRAP_AUTHORIZED_IDS = os.getenv("AUTHORIZED_USER_IDS", "")

access_data = {
    "authorized_users": {ADMIN_ID},
    "pending_users": {},
    "rejected_users": set(),
    "profiles": {}
}


def _now_iso():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _parse_bootstrap_authorized_ids():
    result = set()
    raw = (BOOTSTRAP_AUTHORIZED_IDS or "").strip()
    if not raw:
        return result

    for part in re.split(r"[,;\s]+", raw):
        part = part.strip()
        if part.isdigit():
            result.add(int(part))
    return result


def _default_profile(user_id, first_name="", username=""):
    return {
        "id": int(user_id),
        "first_name": first_name or "",
        "username": username or "",
        "approved_at": _now_iso(),
        "last_seen": None,
        "use_count": 0,
        "last_command": ""
    }


def _normalize_profiles(raw_profiles):
    normalized = {}
    raw_profiles = raw_profiles or {}

    if isinstance(raw_profiles, list):
        for item in raw_profiles:
            if isinstance(item, dict) and str(item.get("id", "")).isdigit():
                uid = str(int(item["id"]))
                normalized[uid] = _default_profile(
                    uid,
                    item.get("first_name", ""),
                    item.get("username", "")
                )
                normalized[uid].update(item)
        return normalized

    if isinstance(raw_profiles, dict):
        for uid, item in raw_profiles.items():
            uid_str = str(uid)
            if not uid_str.isdigit():
                continue
            base = _default_profile(uid_str)
            if isinstance(item, dict):
                base.update(item)
            normalized[uid_str] = base
    return normalized


def _ensure_authorized_profiles():
    for uid in list(access_data.get("authorized_users", set())):
        uid_str = str(int(uid))
        if uid_str not in access_data["profiles"]:
            access_data["profiles"][uid_str] = _default_profile(uid)


def _upsert_profile(user=None, user_id=None, approved=False, command_name=None):
    if user is not None:
        uid = int(user.id)
        first_name = user.first_name or ""
        username = user.username or ""
    else:
        uid = int(user_id)
        first_name = ""
        username = ""

    uid_str = str(uid)
    profile = access_data["profiles"].get(uid_str, _default_profile(uid, first_name, username))

    if first_name:
        profile["first_name"] = first_name
    if username:
        profile["username"] = username

    if approved and not profile.get("approved_at"):
        profile["approved_at"] = _now_iso()

    if command_name is not None:
        profile["last_seen"] = _now_iso()
        profile["last_command"] = command_name
        profile["use_count"] = int(profile.get("use_count", 0) or 0) + 1

    access_data["profiles"][uid_str] = profile
    return profile


def save_access_data():
    try:
        _ensure_authorized_profiles()
        payload = {
            "authorized_users": sorted(int(x) for x in access_data.get("authorized_users", set())),
            "pending_users": access_data.get("pending_users", {}),
            "rejected_users": sorted(int(x) for x in access_data.get("rejected_users", set())),
            "profiles": access_data.get("profiles", {}),
        }

        tmp_fd, tmp_path = tempfile.mkstemp(prefix="access_data_", suffix=".json")
        os.close(tmp_fd)
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

        if os.path.exists(ACCESS_DATA_FILE):
            shutil.copyfile(ACCESS_DATA_FILE, ACCESS_DATA_BACKUP_FILE)

        os.replace(tmp_path, ACCESS_DATA_FILE)
        shutil.copyfile(ACCESS_DATA_FILE, ACCESS_DATA_BACKUP_FILE)
    except Exception as e:
        print(f"ERRORE save_access_data: {e}")


def load_access_data():
    global access_data

    bootstrap_ids = _parse_bootstrap_authorized_ids()
    default_data = {
        "authorized_users": {ADMIN_ID, *bootstrap_ids},
        "pending_users": {},
        "rejected_users": set(),
        "profiles": {}
    }

    for uid in default_data["authorized_users"]:
        default_data["profiles"][str(uid)] = _default_profile(uid)

    source_path = None
    if os.path.exists(ACCESS_DATA_FILE):
        source_path = ACCESS_DATA_FILE
    elif os.path.exists(ACCESS_DATA_BACKUP_FILE):
        source_path = ACCESS_DATA_BACKUP_FILE

    if not source_path:
        access_data = default_data
        save_access_data()
        return

    try:
        with open(source_path, "r", encoding="utf-8") as f:
            payload = json.load(f)

        loaded_authorized = {int(x) for x in payload.get("authorized_users", []) if str(x).isdigit()}
        loaded_authorized.update(default_data["authorized_users"])

        access_data = {
            "authorized_users": loaded_authorized,
            "pending_users": payload.get("pending_users", {}) or {},
            "rejected_users": {int(x) for x in payload.get("rejected_users", []) if str(x).isdigit()},
            "profiles": _normalize_profiles(payload.get("profiles", {}))
        }

        _ensure_authorized_profiles()
        save_access_data()
    except Exception as e:
        print(f"ERRORE load_access_data: {e}")
        access_data = default_data
        save_access_data()


def is_admin(user_id):
    return user_id == ADMIN_ID


def is_authorized(user_id):
    return user_id in access_data["authorized_users"]


def is_pending(user_id):
    return str(user_id) in access_data["pending_users"]


def add_pending(user):
    uid = str(user.id)
    if uid not in access_data["pending_users"]:
        access_data["pending_users"][uid] = {
            "id": user.id,
            "first_name": user.first_name or "",
            "username": user.username or "",
            "requested_at": _now_iso(),
        }
        save_access_data()


def approve_user(user_id):
    uid = int(user_id)
    uid_str = str(uid)
    pending = access_data["pending_users"].get(uid_str, {})
    first_name = pending.get("first_name", "")
    username = pending.get("username", "")

    if uid_str in access_data["pending_users"]:
        del access_data["pending_users"][uid_str]

    access_data["authorized_users"].add(uid)
    access_data["rejected_users"].discard(uid)

    profile = access_data["profiles"].get(uid_str, _default_profile(uid, first_name, username))
    if first_name and not profile.get("first_name"):
        profile["first_name"] = first_name
    if username and not profile.get("username"):
        profile["username"] = username
    if not profile.get("approved_at"):
        profile["approved_at"] = _now_iso()
    access_data["profiles"][uid_str] = profile

    save_access_data()


def reject_user(user_id):
    uid = int(user_id)
    uid_str = str(uid)
    if uid_str in access_data["pending_users"]:
        del access_data["pending_users"][uid_str]
    access_data["rejected_users"].add(uid)
    access_data["authorized_users"].discard(uid)
    save_access_data()


def revoke_user(user_id):
    uid = int(user_id)
    if uid == ADMIN_ID:
        return False
    access_data["authorized_users"].discard(uid)
    save_access_data()
    return True


def track_authorized_usage(user, command_name="messaggio"):
    uid = int(user.id)
    if not (is_admin(uid) or is_authorized(uid)):
        return
    _upsert_profile(user=user, approved=True, command_name=command_name)
    save_access_data()


def format_authorized_users_lines(include_stats=False):
    lines = ["Utenti autorizzati:"]
    profiles = access_data.get("profiles", {})

    for uid in sorted(access_data["authorized_users"]):
        uid_str = str(uid)
        p = profiles.get(uid_str, {})
        name = p.get("first_name") or "-"
        username = f"@{p.get('username')}" if p.get("username") else "-"
        suffix = " (admin)" if uid == ADMIN_ID else ""

        row = f"- ID {uid}{suffix} | {name} | {username}"
        if include_stats:
            approved_at = p.get("approved_at") or "-"
            last_seen = p.get("last_seen") or "-"
            use_count = p.get("use_count", 0)
            last_command = p.get("last_command") or "-"
            row += f" | approvato: {approved_at} | ultimo accesso: {last_seen} | usi: {use_count} | ultimo comando: {last_command}"
        lines.append(row)

    return lines

def send_welcome_media(chat_id):
    if not WELCOME_MEDIA_ENABLED:
        return

    try:
        if os.path.exists(WELCOME_MEDIA_PATH):
            with open(WELCOME_MEDIA_PATH, "rb") as media_file:
                try:
                    bot.send_video(chat_id, media_file, supports_streaming=True)
                except Exception:
                    media_file.seek(0)
                    bot.send_animation(chat_id, media_file)
            return

        if WELCOME_MEDIA_URL:
            try:
                bot.send_video(chat_id, WELCOME_MEDIA_URL, supports_streaming=True)
            except Exception:
                bot.send_animation(chat_id, WELCOME_MEDIA_URL)
    except Exception:
        pass

def request_access_text():
    return (
        "Benvenuto in NCC Sanzioni Bot.\n\n"
        "Questo assistente supporta l’analisi preliminare delle possibili violazioni in materia NCC, con indicazione di:\n"
        "- riferimento normativo\n"
        "- voce operativa\n"
        "- importi\n"
        "- sanzioni accessorie\n"
        "- dicitura verbale\n"
        "- verifiche finali\n\n"
        "Uso interno-operativo.\n"
        "Le risultanze vanno sempre verificate su normativa vigente, prontuario e disposizioni di servizio.\n\n"
        "Il tuo accesso è in attesa di approvazione amministratore."
    )

def authorized_start_text(user_id):
    text = (
        f"Benvenuto in NCC Sanzioni Bot.\n\n"
        f"Accesso autorizzato.\n"
        f"Il tuo user id è: {user_id}\n\n"
        "Comandi disponibili:\n"
        "/controllo - checklist documentale guidata\n/caso - descrivi liberamente il fatto\n"
        "/checklist - controlli operativi\n"
        "/documenti - documenti da controllare\n"
        "/norme - riferimenti principali\n"
        "/targa - verifica targa in archivio NCC\n"
        "/riattiva - istruzioni per riattivare il servizio\n"
        "/reset - annulla caso in corso"
    )

    if user_id == ADMIN_ID:
        text += (
            "\n/restartbot - riavvia il servizio Render\n"
            "/deploybot - avvia un deploy Render"
        )

    return text

def notify_admin_new_request(user):
    username_line = f"@{user.username}" if user.username else "-"
    text = (
        "Richiesta accesso al bot\n\n"
        f"Nome: {user.first_name or '-'}\n"
        f"Username: {username_line}\n"
        f"ID: {user.id}\n\n"
        f"Per approvare:\n/approva {user.id}\n\n"
        f"Per rifiutare:\n/rifiuta {user.id}"
    )
    bot.send_message(ADMIN_ID, text)

def notify_admin_missing_plate(user, plate):
    username_line = f"@{user.username}" if user.username else "-"
    text = (
        "RICHIESTA CENSIMENTO TARGA\n\n"
        f"Targa: {plate}\n"
        f"Richiesta da: {user.first_name or '-'}\n"
        f"Username: {username_line}\n"
        f"User ID: {user.id}\n\n"
        "Verificare e, se confermata, censire la targa nell'archivio."
    )
    bot.send_message(ADMIN_ID, text)

def ensure_authorized(message):
    uid = message.from_user.id
    if is_admin(uid) or is_authorized(uid):
        command_name = (message.text or "messaggio").strip()[:100] or "messaggio"
        track_authorized_usage(message.from_user, command_name)
        return True

    send_welcome_media(message.chat.id)

    if is_pending(uid):
        bot.reply_to(
            message,
            "Accesso non ancora autorizzato.\nLa tua richiesta è già stata inviata all'amministratore."
        )
        return False

    add_pending(message.from_user)
    notify_admin_new_request(message.from_user)
    bot.reply_to(message, request_access_text())
    return False

def render_headers():
    if not RENDER_API_KEY:
        return None
    return {
        "Authorization": f"Bearer {RENDER_API_KEY}",
        "Accept": "application/json"
    }

def restart_render_service():
    if not RENDER_API_KEY or not RENDER_SERVICE_ID:
        return False, "Variabili Render mancanti: RENDER_API_KEY o RENDER_SERVICE_ID."

    url = f"https://api.render.com/v1/services/{RENDER_SERVICE_ID}/restart"
    try:
        response = requests.post(url, headers=render_headers(), timeout=20)
        if response.status_code in [200, 202]:
            return True, "Riavvio del servizio richiesto correttamente."
        return False, f"Errore Render restart: HTTP {response.status_code} - {response.text}"
    except Exception as e:
        return False, f"Eccezione restart Render: {e}"

def deploy_render_service():
    if not RENDER_API_KEY or not RENDER_SERVICE_ID:
        return False, "Variabili Render mancanti: RENDER_API_KEY o RENDER_SERVICE_ID."

    url = f"https://api.render.com/v1/services/{RENDER_SERVICE_ID}/deploys"
    try:
        response = requests.post(url, headers=render_headers(), timeout=20)
        if response.status_code in [200, 201]:
            return True, "Deploy del servizio richiesto correttamente."
        return False, f"Errore Render deploy: HTTP {response.status_code} - {response.text}"
    except Exception as e:
        return False, f"Eccezione deploy Render: {e}"

def public_wake_up_message():
    return (
        "Richiesta di riattivazione inviata.\n\n"
        "Il servizio gratuito può impiegare circa 30-60 secondi per tornare operativo.\n\n"
        f"Apri questo link per favorire il risveglio del servizio:\n{RENDER_PUBLIC_URL}\n\n"
        "Attendi circa 1 minuto e poi riprova con /start oppure con il comando che stavi usando.\n\n"
        "Se il bot non si riattiva, contatta l'amministratore per il riavvio del servizio."
    )

# =========================
# STATO CONVERSAZIONI
# =========================

user_states = {}


def _serialize_for_json(obj):
    if isinstance(obj, dict):
        return {str(k): _serialize_for_json(v) for k, v in obj.items()}
    if isinstance(obj, set):
        return [_serialize_for_json(x) for x in obj]
    if isinstance(obj, list):
        return [_serialize_for_json(x) for x in obj]
    return obj


def save_user_states():
    try:
        payload = {str(k): _serialize_for_json(v) for k, v in user_states.items()}
        with open(USER_STATES_FILE, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"ERRORE save_user_states: {e}")


def load_user_states():
    global user_states
    try:
        if not os.path.exists(USER_STATES_FILE):
            return
        with open(USER_STATES_FILE, "r", encoding="utf-8") as f:
            payload = json.load(f)
        user_states = {int(k): v for k, v in (payload or {}).items()}
    except Exception as e:
        print(f"ERRORE load_user_states: {e}")
        user_states = {}

load_access_data()
load_user_states()

# mode:
# - free_case: attesa descrizione libera
# - clarification: attesa risposta a domanda integrativa

# =========================
# FUNZIONI FORMATO
# =========================

def clear_case(chat_id):
    if chat_id in user_states:
        del user_states[chat_id]
        save_user_states()

def get_state(chat_id):
    return user_states.get(chat_id)


def normalize_article_key(article_key):
    if not article_key:
        return None
    key = str(article_key).strip().lower().replace('/', '')
    aliases = {
        'art85': 'art85',
        '85': 'art85',
        'art116': 'art116',
        '116': 'art116',
        'art3l21': 'art3l21',
        'art3': 'art3l21',
        '3l21': 'art3l21',
        'art11l21': 'art11l21',
        'art11': 'art11l21',
        '11l21': 'art11l21',
        'art180': 'art180',
        '180': 'art180',
        'art126': 'art126',
        '126': 'art126',
    }
    return aliases.get(key)


def format_articolo(article_key):
    key = normalize_article_key(article_key)
    item = ARTICOLI_DB.get(key) if key else None
    if not item:
        return (
            'Articolo non disponibile nel database interno.\n\n'
            'Usa uno di questi comandi: /art85, /art116, /art3l21, /art11l21'
        )

    lines = [item['titolo'], '', item['testo']]
    link = item.get('link')
    if link:
        lines.extend(['', f'Fonte normativa / consultazione: {link}'])
    return '\n'.join(lines)


def get_article_keys_for_result(main_code=None, concurrent_codes=None):
    concurrent_codes = concurrent_codes or []
    article_keys = []

    def add(key):
        norm = normalize_article_key(key)
        if norm and norm not in article_keys:
            article_keys.append(norm)

    code_to_article = {
        '085-02': 'art85',
        '085-04': 'art85',
        '085-05': 'art85',
        '085-06': 'art85',
        '085-07': 'art85',
        '085-08': 'art85',
        '085-09': 'art85',
        '116-01': 'art116',
        '116-02': 'art116',
        '116-03': 'art116',
        '116-04': 'art116',
        '116-05': 'art116',
        '116-06': 'art116',
        '180-01': 'art180',
        '180-DOC': 'art180',
        '180-01DOC': 'art180',
        '180-03': 'art180',
        '180-06': 'art180',
        '180-09': 'art180',
        'CDS_126_11': 'art126',
    }

    add(code_to_article.get(main_code))

    if main_code in {'085-05', '085-06', '085-07', '085-08'}:
        add('art3l21')
        add('art11l21')

    for code in concurrent_codes:
        add(code_to_article.get(code))

    return article_keys


def build_article_markup(article_keys=None):
    article_keys = article_keys or []
    article_keys = [normalize_article_key(k) for k in article_keys]
    article_keys = [k for k in article_keys if k]
    if not article_keys:
        return None

    label_map = {
        'art85': 'Art. 85 CdS',
        'art116': 'Art. 116 CdS',
        'art3l21': 'Art. 3 L. 21/1992',
        'art11l21': 'Art. 11 L. 21/1992',
        'art180': 'Art. 180 CdS',
        'art126': 'Art. 126 CdS',
    }

    markup = types.InlineKeyboardMarkup(row_width=2)
    buttons = []
    for key in article_keys:
        label = label_map.get(key, key)
        buttons.append(types.InlineKeyboardButton(label, callback_data=f'article:{key}'))
    if buttons:
        markup.add(*buttons)
    return markup

def build_port_common_cases_markup():
    markup = types.InlineKeyboardMarkup(row_width=1)
    markup.add(
        types.InlineKeyboardButton(
            "KB presente + mezzo uso proprio",
            callback_data="porto_case:uso_proprio_kb"
        ),
        types.InlineKeyboardButton(
            "Abusivo totale",
            callback_data="porto_case:abusivo_totale"
        ),
        types.InlineKeyboardButton(
            "NCC con licenza/KB ma senza prenotazione",
            callback_data="porto_case:procacciamento"
        ),
        types.InlineKeyboardButton(
            "Altro caso porto",
            callback_data="porto_case:altro"
        )
    )
    return markup

def build_plate_not_found_markup(plate):
    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton(
            "SEGNALA PER CENSIMENTO",
            callback_data=f"plate_report:{plate}"
        )
    )
    return markup

def build_main_menu():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add(
        types.KeyboardButton("Inserisci un caso NCC"),
        types.KeyboardButton("Controlli operativi")
    )
    kb.add(
        types.KeyboardButton("Checklist documentale"),
        types.KeyboardButton("Documenti da controllare")
    )
    kb.add(
        types.KeyboardButton("Verifica targa"),
        types.KeyboardButton("Norme principali")
    )
    kb.add(
        types.KeyboardButton("Casi comuni porto"),
        types.KeyboardButton("Reset")
    )
    kb.add(
        types.KeyboardButton("Help")
    )
    return kb

def build_pdf_markup(main_code=None, concurrent_codes=None, procedural_flags=None):
    concurrent_codes = _dedupe_keep_order(concurrent_codes or [])
    procedural_flags = procedural_flags or {}

    markup = types.InlineKeyboardMarkup(row_width=2)
    added = False

    def add_button(code, label):
        nonlocal added
        if code and PDF_MODELS.get(code):
            markup.add(types.InlineKeyboardButton(label, url=PDF_MODELS[code]))
            added = True

    add_button(main_code, "PDF Verbale principale")

    for idx, code in enumerate(concurrent_codes, start=2):
        add_button(code, f"PDF Verbale {idx}")

    accessory_actions = _build_accessory_actions(main_code, concurrent_codes)
    communications = _build_communications(procedural_flags)

    if "Sequestro/confisca veicolo" in accessory_actions:
        add_button("SEQUESTRO_85", "PDF Sequestro")
    if "Fermo veicolo 60 giorni" in accessory_actions:
        add_button("FERMO_116", "PDF Fermo")

    if "Prefetto" in communications:
        add_button("COM_PREFETTO", "PDF Comunicazione Prefetto")
    if "UMC" in communications:
        add_button("COM_UMC", "PDF Comunicazione UMC")
    if "Comune / ente rilasciante" in communications:
        add_button("COM_COMUNE", "PDF Comunicazione Comune")
    if "Segnalazione RENT" in communications:
        add_button("COM_RENT", "PDF Segnalazione RENT")
    if "Segnalazione ruolo/albo conducenti" in communications:
        add_button("COM_RUOLO", "PDF Segnalazione Ruolo")

    return markup if added else None        

def build_final_result_markup(payload):
    if not payload:
        return None

    markup = types.InlineKeyboardMarkup(row_width=2)

    markup.add(
        types.InlineKeyboardButton("ESITO RAPIDO", callback_data="final:quick"),
        types.InlineKeyboardButton("ATTI ACCESSORI", callback_data="final:accessori"),
    )

    verbali = payload.get("verbali", [])
    if len(verbali) >= 1:
        markup.add(types.InlineKeyboardButton("VERBALE 1", callback_data="final:v1"))
    if len(verbali) >= 2:
        markup.add(types.InlineKeyboardButton("VERBALE 2", callback_data="final:v2"))
    if len(verbali) >= 3:
        markup.add(types.InlineKeyboardButton("VERBALE 3", callback_data="final:v3"))
    if len(verbali) >= 4:
        markup.add(types.InlineKeyboardButton("VERBALE 4", callback_data="final:v4"))

    markup.add(
        types.InlineKeyboardButton("COMUNICAZIONI", callback_data="final:comunicazioni"),
        types.InlineKeyboardButton("ARTICOLI", callback_data="final:articoli"),
    )

    return markup

def build_quick_payload_from_codes(main_code, concurrent_codes=None, extra_articles=None):
    concurrent_codes = concurrent_codes or []
    extra_articles = extra_articles or []

    main = VIOLATIONS.get(main_code, {})
    concurrent = [VIOLATIONS[c] for c in concurrent_codes if c in VIOLATIONS]

    quick_parts = []
    if main:
        quick_parts.append(format_compact_violation(main_code))

    for code in concurrent_codes:
        if code in VIOLATIONS:
            quick_parts.append("\n\n" + format_compact_violation(code))

    article_keys = get_article_keys_for_result(main_code, concurrent_codes)
    for k in extra_articles:
        nk = normalize_article_key(k)
        if nk and nk not in article_keys:
            article_keys.append(nk)

    articoli_text = "\n\n".join(format_articolo(k) for k in article_keys) if article_keys else "Articoli non disponibili."

    accessori = []
    if main.get("accessories"):
        accessori.extend(main["accessories"])
    for item in concurrent:
        accessori.extend(item.get("accessories", []))

    accessori = list(dict.fromkeys(accessori))
    accessori_text = "\n".join(f"- {x}" for x in accessori) if accessori else "Nessuna sanzione accessoria indicata."

    verbali = []
    if main.get("verbal_text"):
        verbali.append(main["verbal_text"])
    for item in concurrent:
        if item.get("verbal_text"):
            verbali.append(item["verbal_text"])

    comunicazioni = []
    if main_code in {"085-02"}:
        comunicazioni.append("- Trasmissione al Prefetto entro 10 giorni.")
    if main_code in {"085-04"}:
        comunicazioni.append("- Comunicazione al Prefetto entro 5 giorni per i presupposti della revoca.")
    if main_code in {"085-05", "085-06", "085-07", "085-08"}:
        comunicazioni.append("- Trasmissione documento ritirato all'UMC competente.")
    if "116-06" in concurrent_codes:
        comunicazioni.append("- Valutare fermo del veicolo per 60 giorni.")

    payload = {
        "quick": "\n\n".join(quick_parts) if quick_parts else "Esito non disponibile.",
        "verbali": verbali,
        "accessori": accessori_text,
        "comunicazioni": "\n".join(comunicazioni) if comunicazioni else "Nessuna comunicazione principale preimpostata.",
        "articoli": articoli_text,
    }
    return payload
    
def get_question_buttons(question_key):
    mapping = {
        'vehicle_authorized': [('SI', 'si'), ('NO', 'no')],
        'service_to_third': [('SI', 'si'), ('NO', 'no'), ('DUBBIO', 'dubbio')],
        'separate_payment': [('SI', 'si'), ('NO', 'no')],
        'kb': [('SI', 'si'), ('NO', 'no')],
        'patente_idonea': [('SI', 'si'), ('NO', 'no')],
        'rent_registered': [('SI', 'si'), ('NO', 'no'), ('NON VERIF.', 'non_verificato')],
        'ruolo_conducenti': [('SI', 'si'), ('NO', 'no'), ('NON VERIF.', 'non_verificato')],
        'incauto_affidamento': [('SI', 'si'), ('NO', 'no')],
        'public_waiting': [('SI', 'si'), ('NO', 'no')],
        'taxi_commune': [('SI', 'si'), ('NO', 'no')],
        'booking': [('SI', 'si'), ('NO', 'no')],
        'violation_type': [('ART. 3/11', 'art3_11'), ('ALTRE PRESCR.', 'other_auth'), ('NON CHIARO', 'none')],
        'foglio_status': [('PRESENTE', 'presente'), ('ASSENTE', 'assente'), ('IRREGOLARE', 'irregolare'), ('NON ESIBITO', 'non_esibito')],
        'recurrence': [('PRIMA', 'first'), ('2^ QUINQ.', '2_5y'), ('3^ QUINQ.', '3_5y'), ('4+', '4plus_5y')],
        'recurrence_triennio': [('PRIMA', 'first'), ('2^ TRIENNIO', 'second_3y')],
        'control_patente_status': [('VALIDA', 'valida'), ('SCADUTA', 'scaduta'), ('NON IDONEA', 'non_idonea'), ('NON ESIBITA', 'non_esibita')],
        'control_kb_status': [('VALIDO', 'valido'), ('SCADUTO', 'scaduto'), ('NON IDONEO/MAI', 'non_idoneo'), ('NON ESIBITO', 'non_esibito'), ('NON DOVUTO', 'non_dovuto')],
        'control_autorizzazione_status': [('REGOLARE', 'regolare'), ('NON ESIBITA', 'non_esibita'), ('NON AUTORIZZATO', 'non_autorizzato')],
        'control_foglio_status': [('REGOLARE', 'regolare'), ('IRREGOLARE', 'irregolare'), ('ASSENTE', 'assente'), ('NON ESIBITO', 'non_esibito')],
        'control_rent_status': [('REGOLARE', 'si'), ('NON REGOLARE', 'no'), ('NON VERIF.', 'non_verificato')],
        'control_ruolo_status': [('REGOLARE', 'si'), ('ASSENTE', 'no'), ('NON VERIF.', 'non_verificato')],
        'control_owner_type': [('PERSONA FISICA', 'persona_fisica'), ('COOP./SRL', 'cooperativa_srl'), ('AGENZIA VIAGGI', 'agenzia_viaggi'), ('ALTRO', 'altro')],
        'control_circulation_use': [('USO TERZI/NCC', 'uso_terzi_ncc'), ('USO PROPRIO', 'uso_proprio'), ('NON LETTO', 'non_letto')],
        'control_trip_nature': [('NCC PURO', 'ncc_puro'), ('PACCHETTO AGENZIA', 'agenzia_pacchetto'), ('NAVETTA ACCESS.', 'navetta_accessoria'), ('DUBBIO', 'dubbio')],
    }
    return mapping.get(question_key, [])


def build_combined_markup(article_keys=None, question_key=None, force_ctrl_answer=False):
    article_keys = article_keys or []
    q_buttons = get_question_buttons(question_key) if question_key else []
    if not article_keys and not q_buttons:
        return None

    markup = types.InlineKeyboardMarkup(row_width=2)

    if article_keys:
        label_map = {
            'art85': 'Art. 85 CdS',
            'art116': 'Art. 116 CdS',
            'art3l21': 'Art. 3 L. 21/1992',
            'art11l21': 'Art. 11 L. 21/1992',
            'art180': 'Art. 180 CdS',
            'art126': 'Art. 126 CdS',
        }
        article_buttons = []
        for key in article_keys:
            label = label_map.get(key, key)
            article_buttons.append(types.InlineKeyboardButton(label, callback_data=f'article:{key}'))
        if article_buttons:
            markup.add(*article_buttons)

    if q_buttons:
        row = []
        for label, value in q_buttons:
            if force_ctrl_answer:
                prefix = 'ctrl_answer'
            else:
                prefix = 'ctrl_answer' if (question_key or '').startswith('control_') else 'answer'

            row.append(types.InlineKeyboardButton(label, callback_data=f'{prefix}:{value}'))
            if len(row) == 2:
                markup.row(*row)
                row = []
        if row:
            markup.row(*row)

    return markup



def infer_article_keys_from_text(text):
    if not text:
        return []
    mapping = {
        '/art85': 'art85',
        '/art116': 'art116',
        '/art3l21': 'art3l21',
        '/art11l21': 'art11l21',
        '/art180': 'art180',
        '/art126': 'art126',
    }
    found = []
    lowered = str(text).lower()
    for cmd, key in mapping.items():
        if cmd in lowered and key not in found:
            found.append(key)
    return found


def reply_with_article_buttons(message, text, article_keys=None, disable_web_page_preview=True):
    keys = article_keys or infer_article_keys_from_text(text)
    state = get_state(message.chat.id)
    question_key = None
    if state and state.get('mode') == 'clarification' and state.get('pending_question'):
        question_key = state['pending_question'].get('key')
    markup = build_combined_markup(keys, question_key=question_key)
    bot.reply_to(message, text, reply_markup=markup, disable_web_page_preview=disable_web_page_preview)



def article_shortcuts_from_result(main_code=None, concurrent_codes=None):
    article_keys = get_article_keys_for_result(main_code, concurrent_codes)
    if not article_keys:
        return ''

    label_map = {
        'art85': '/art85',
        'art116': '/art116',
        'art3l21': '/art3l21',
        'art11l21': '/art11l21',
        'art180': '/art180',
        'art126': '/art126',
    }
    commands = [label_map[k] for k in article_keys if k in label_map]
    if not commands:
        return ''

    return 'ARTICOLI RICHIAMABILI\n' + '\n'.join(f'- {cmd}' for cmd in commands)




def normalize_violation_code(code):
    if not code:
        return None
    key = str(code).strip().upper()
    if key in VIOLATIONS:
        return key
    compact = re.sub(r"[^A-Z0-9]", "", key)
    aliases = {
        "08502": "085-02",
        "08504": "085-04",
        "08505": "085-05",
        "08506": "085-06",
        "08507": "085-07",
        "08508": "085-08",
        "08509": "085-09",
        "11606": "116-06",
        "18001": "180-01",
        "180DOC": "180-DOC",
        "18001DOC": "180-01DOC",
        "18003": "180-03",
        "18006": "180-06",
        "18009": "180-09",
    }
    return aliases.get(compact)


def build_violation_markup_for_article(article_key):
    key = normalize_article_key(article_key)
    article_to_codes = {
        "art85": ["085-02", "085-04", "085-05", "085-06", "085-07", "085-08", "085-09"],
        "art116": ["116-06"],
        "art3l21": ["085-05", "085-06", "085-07", "085-08"],
        "art11l21": ["085-05", "085-06", "085-07", "085-08"],
        "art180": ["180-01", "180-DOC", "180-01DOC", "180-03", "180-06", "180-09"],
        "art126": [],
    }
    codes = [c for c in article_to_codes.get(key, []) if c in VIOLATIONS]
    if not codes:
        return None
    markup = types.InlineKeyboardMarkup(row_width=2)
    buttons = []
    for code in codes:
        label = code
        buttons.append(types.InlineKeyboardButton(label, callback_data=f"viol:{code}"))
    for i in range(0, len(buttons), 2):
        markup.row(*buttons[i:i+2])
    return markup


def send_long_message(chat_id, text, reply_markup=None, disable_web_page_preview=True, chunk_size=3500):
    text = text or ""
    chunks = []

    while len(text) > chunk_size:
        split_at = text.rfind("\n", 0, chunk_size)
        if split_at == -1:
            split_at = chunk_size
        chunks.append(text[:split_at])
        text = text[split_at:].lstrip("\n")

    if text:
        chunks.append(text)

    if not chunks:
        chunks = [""]

    for i, chunk in enumerate(chunks):
        markup = reply_markup if i == 0 else None
        bot.send_message(
            chat_id,
            chunk,
            reply_markup=markup,
            disable_web_page_preview=disable_web_page_preview
        )

PDF_MODELS = {
    "085-02": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/085-02_art85_c4_prima_violazione.pdf",
    "085-04": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/085-02_art85_c4_seconda_nel_triennio.pdf",
    "085-05": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/085-05_art85_c4bis_prima_nel_quinquennio.pdf",
    "085-06": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/085-05_art85_c4bis_seconda_nel_quinquennio.pdf",
    "085-07": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/085-05_art85_c4bis_terza_nel_quinquennio.pdf",
    "085-08": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/085-05_art85_c4bis_quarta_o_successiva_nel_quinquennio.pdf",
    "085-09": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/085-09_art85_c4ter_altre_prescrizioni.pdf",
    "116-06": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/116-06_art116_senza_kb_cqc.pdf",
    "180-01": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/180-01_foglio_di_servizio_non_esibito.pdf",
    "180-DOC": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/180-01DOC_patente_o_carta_non_al_seguito.pdf",
    "180-01DOC": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/180-01DOC_patente_o_carta_non_al_seguito.pdf",
    "180-03": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/180-03_certificato_assicurativo_non_al_seguito.pdf",
    "180-06": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/180-06_autorizzazione_ncc_non_al_seguito.pdf",
    "180-09": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/180-09_kb_cqc_non_al_seguito.pdf",
    "SEQUESTRO_85": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/VERBALE_SEQUESTRO_CUSTODIA.pdf",
    "FERMO_116": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/VERBALE_FERMO_O_SEQUESTRO.pdf",
    "AVVISO_FERMO": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/AVVISO_FERMO.pdf",
    "COM_PREFETTO": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/COM_PREFETTO.pdf",
    "COM_UMC": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/COM_UMC.pdf",
    "COM_COMUNE": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/COM_COMUNE.pdf",
    "COM_RENT": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/COM_RENT.pdf",
    "COM_RUOLO": "https://raw.githubusercontent.com/lord26300/ncc-sanzioni-bot/main/pdf_templates/COM_RUOLO.pdf",
}

def _short_violation_line(code):
    v = VIOLATIONS.get(code)
    if not v:
        return code
    return f"{code} | {v['article']} | {v['title']}"

def _dedupe_keep_order(items):
    out = []
    for item in items or []:
        if item and item not in out:
            out.append(item)
    return out

def _build_accessory_actions(main_code, concurrent_codes=None):
    concurrent_codes = concurrent_codes or []
    actions = []

    def add(text):
        if text and text not in actions:
            actions.append(text)

    if main_code in {"085-02", "085-04"}:
        add("Sequestro/confisca veicolo")
    if main_code in {"085-05", "085-06", "085-07", "085-08"}:
        add("Ritiro/sospensione documento di circolazione")
    if "116-06" in concurrent_codes:
        add("Fermo veicolo 60 giorni")

    return actions

def _build_communications(procedural_flags=None):
    procedural_flags = procedural_flags or {}
    signals = _dedupe_keep_order(procedural_flags.get("segnalazioni", []))
    mapped = []

    for item in signals:
        low = item.lower()
        if "prefett" in low:
            mapped.append("Prefetto")
        elif "umc" in low:
            mapped.append("UMC")
        elif "comune" in low or "ente rilasciante" in low:
            mapped.append("Comune / ente rilasciante")
        elif "rent" in low:
            mapped.append("Segnalazione RENT")
        elif "ruolo" in low or "albo" in low:
            mapped.append("Segnalazione ruolo/albo conducenti")
        else:
            mapped.append(item)

    return _dedupe_keep_order(mapped)

def _compact_details_for_code(code):
    v = VIOLATIONS[code]
    lines = []
    lines.append(f"{code} — {v['article']}")
    lines.append(v["title"])
    lines.append(f"Sanzione: {v['edictal']}")
    lines.append(f"PMR: {v['pmr']}")
    if v.get("accessories"):
        lines.append("Accessorie:")
        for a in v["accessories"]:
            lines.append(f"- {a}")
    if v.get("short_ready_text"):
        lines.append("Testo sintetico:")
        lines.append(v["short_ready_text"])
    if v.get("fields_to_fill"):
        lines.append("Dati da completare:")
        for f in v["fields_to_fill"]:
            lines.append(f"- {f}")
    return "\n".join(lines)

def build_quick_summary(main_code, concurrent_codes=None, procedural_flags=None, ancillary_findings=None):
    concurrent_codes = _dedupe_keep_order(concurrent_codes or [])
    ancillary_findings = _dedupe_keep_order(ancillary_findings or [])
    accessory_actions = _build_accessory_actions(main_code, concurrent_codes)
    communications = _build_communications(procedural_flags)

    lines = []
    lines.append("ESITO RAPIDO")
    lines.append("")
    lines.append("VERBALI DA FARE")
    lines.append(f"1) {_short_violation_line(main_code)}")
    for idx, code in enumerate(concurrent_codes, start=2):
        lines.append(f"{idx}) {_short_violation_line(code)}")

    if accessory_actions:
        lines.append("")
        lines.append("ATTI ACCESSORI")
        for idx, item in enumerate(accessory_actions, start=1):
            lines.append(f"{idx}) {item}")

    if communications:
        lines.append("")
        lines.append("COMUNICAZIONI")
        for idx, item in enumerate(communications, start=1):
            lines.append(f"{idx}) {item}")

    if ancillary_findings:
        lines.append("")
        lines.append("ALTRE ANOMALIE DA ANNOTARE")
        for idx, item in enumerate(ancillary_findings, start=1):
            lines.append(f"{idx}) {item}")

    return "\n".join(lines)

def format_multiple(main_code, concurrent_codes=None, extra_notes=None, level=None, procedural_flags=None, ancillary_findings=None):
    concurrent_codes = _dedupe_keep_order(concurrent_codes or [])
    extra_notes = _dedupe_keep_order(extra_notes or [])
    procedural_flags = procedural_flags or {}
    ancillary_findings = _dedupe_keep_order(ancillary_findings or [])

    accessory_actions = _build_accessory_actions(main_code, concurrent_codes)
    communications = _build_communications(procedural_flags)

    lines = []
    lines.append(build_quick_summary(
        main_code,
        concurrent_codes=concurrent_codes,
        procedural_flags=procedural_flags,
        ancillary_findings=ancillary_findings,
    ))

    lines.append("")
    lines.append("DETTAGLIO VERBALI")

    all_codes = [main_code] + concurrent_codes
    for idx, code in enumerate(all_codes, start=1):
        lines.append("")
        lines.append(f"VERBALE {idx}")
        lines.append(_compact_details_for_code(code))

    if accessory_actions:
        lines.append("")
        lines.append("DETTAGLIO ATTI ACCESSORI")
        for idx, item in enumerate(accessory_actions, start=1):
            lines.append(f"{idx}) {item}")

    if communications:
        lines.append("")
        lines.append("DETTAGLIO COMUNICAZIONI")
        for idx, item in enumerate(communications, start=1):
            lines.append(f"{idx}) {item}")

    verbale_additions = _dedupe_keep_order(procedural_flags.get("verbale_additions", []))
    if verbale_additions:
        lines.append("")
        lines.append("ANNOTAZIONI DA INSERIRE NEL VERBALE")
        for item in verbale_additions:
            lines.append(f"- {item}")

    if extra_notes:
        lines.append("")
        lines.append("NOTE OPERATIVE")
        for note in extra_notes:
            lines.append(f"- {note}")

    lines.append("")
    lines.append("ARTICOLI RICHIAMABILI")
    shortcuts = article_shortcuts_from_result(main_code, concurrent_codes)
    lines.append(shortcuts if shortcuts else "- Nessuno")

    lines.append("")
    lines.append("AVVERTENZA")
    lines.append("Verificare sempre normativa vigente, prontuario del comando, disciplina locale e dati concreti del caso.")

    return "\n".join(lines)

def format_compact_violation(code):
    v = VIOLATIONS[code]
    lines = []
    lines.append(f"{code} | {v['article']} | {v['title']}")
    lines.append(f"- PMR: {v['pmr']}")
    lines.append(f"- Riduzione 30%: {v['reduced_30']}")
    lines.append(f"- Oltre 60 gg: {v['over_60']}")
    lines.append(f"- Limiti edittali: {v['edictal']}")
    lines.append("- Sanzioni accessorie:")
    for a in v["accessories"]:
        lines.append(f"  • {a}")
    lines.append("- Dicitura verbale:")
    lines.append(f"  {v['verbal_text']}")
    if v.get("fields_to_fill"):
        lines.append("- Dati da completare:")
        for f in v["fields_to_fill"]:
            lines.append(f"  • {f}")
    if v.get("short_ready_text"):
        lines.append("- Verbale sintetico pronto:")
        lines.append(f"  {v['short_ready_text']}")
    return "\n".join(lines)

def build_final_payload(main_code, concurrent_codes=None, extra_notes=None, procedural_flags=None, ancillary_findings=None):
    concurrent_codes = _dedupe_keep_order(concurrent_codes or [])
    extra_notes = _dedupe_keep_order(extra_notes or [])
    procedural_flags = procedural_flags or {}
    ancillary_findings = _dedupe_keep_order(ancillary_findings or [])

    all_codes = [main_code] + concurrent_codes
    accessory_actions = _build_accessory_actions(main_code, concurrent_codes)
    communications = _build_communications(procedural_flags)

    quick_lines = []
    quick_lines.append("ESITO RAPIDO")
    quick_lines.append("")
    quick_lines.append(f"Verbali da fare: {len(all_codes)}")
    quick_lines.append(f"Atti accessori: {len(accessory_actions)}")
    quick_lines.append(f"Comunicazioni: {len(communications)}")
    quick_lines.append("")
    quick_lines.append("VERBALI DA REDIGERE")
    for idx, code in enumerate(all_codes, start=1):
        quick_lines.append(f"{idx}) {_short_violation_line(code)}")

    if accessory_actions:
        quick_lines.append("")
        quick_lines.append("ATTI ACCESSORI")
        for idx, item in enumerate(accessory_actions, start=1):
            quick_lines.append(f"{idx}) {item}")

    if communications:
        quick_lines.append("")
        quick_lines.append("COMUNICAZIONI")
        for idx, item in enumerate(communications, start=1):
            quick_lines.append(f"{idx}) {item}")

    if ancillary_findings:
        quick_lines.append("")
        quick_lines.append("ALTRE ANOMALIE DA ANNOTARE")
        for idx, item in enumerate(ancillary_findings, start=1):
            quick_lines.append(f"{idx}) {item}")

    verbali = []
    for idx, code in enumerate(all_codes, start=1):
        text = []
        text.append(f"VERBALE {idx}")
        text.append("")
        text.append(_compact_details_for_code(code))
        verbali.append("\n".join(text))

    accessori_text = "ATTI ACCESSORI\n\n"
    if accessory_actions:
        for idx, item in enumerate(accessory_actions, start=1):
            accessori_text += f"{idx}) {item}\n"
    else:
        accessori_text += "Nessun atto accessorio.\n"

    comunicazioni_text = "COMUNICAZIONI\n\n"
    if communications:
        for idx, item in enumerate(communications, start=1):
            comunicazioni_text += f"{idx}) {item}\n"
    else:
        comunicazioni_text += "Nessuna comunicazione conseguente.\n"

    verbale_additions = _dedupe_keep_order(procedural_flags.get("verbale_additions", []))
    if verbale_additions:
        comunicazioni_text += "\nANNOTAZIONI DA INSERIRE NEL VERBALE\n"
        for item in verbale_additions:
            comunicazioni_text += f"- {item}\n"

    if extra_notes:
        comunicazioni_text += "\nNOTE OPERATIVE\n"
        for note in extra_notes:
            comunicazioni_text += f"- {note}\n"

    articoli_text = "ARTICOLI RICHIAMABILI\n\n"
    shortcuts = article_shortcuts_from_result(main_code, concurrent_codes)
    articoli_text += shortcuts if shortcuts else "- Nessuno"

    return {
        "quick": "\n".join(quick_lines),
        "verbali": verbali,
        "accessori": accessori_text.strip(),
        "comunicazioni": comunicazioni_text.strip(),
        "articoli": articoli_text.strip(),
        "main_code": main_code,
        "concurrent_codes": concurrent_codes,
        "procedural_flags": procedural_flags,
    }

def format_multiple(main_code, concurrent_codes=None, extra_notes=None, level=None, procedural_flags=None, ancillary_findings=None):
    if concurrent_codes is None:
        concurrent_codes = []
    if extra_notes is None:
        extra_notes = []
    if procedural_flags is None:
        procedural_flags = {}
    if ancillary_findings is None:
        ancillary_findings = []

    v = VIOLATIONS[main_code]
    livello = level if level else "non specificato"

    lines = []
    lines.append("ESITO FINALE")
    lines.append(f"Livello di affidabilità: {livello}")
    lines.append(v["title"])
    lines.append("")
    lines.append("RIFERIMENTO")
    lines.append(v["article"])
    lines.append("")
    lines.append("VOCE OPERATIVA")
    lines.append(main_code)
    lines.append("")
    lines.append("IMPORTI")
    lines.append(f"- Pagamento in misura ridotta: {v['pmr']}")
    lines.append(f"- Riduzione 30% entro 5 gg: {v['reduced_30']}")
    lines.append(f"- Pagamento oltre 60 gg: {v['over_60']}")
    lines.append(f"- Limiti edittali: {v['edictal']}")
    lines.append("")
    lines.append("SANZIONI ACCESSORIE")
    for a in v["accessories"]:
        lines.append(f"- {a}")
    lines.append("")
    lines.append("DICITURA VERBALE")
    lines.append(v["verbal_text"])

    if procedural_flags.get("verbale_additions"):
        lines.append("")
        lines.append("INTEGRAZIONE DA INSERIRE NEL VERBALE")
        for item in procedural_flags["verbale_additions"]:
            lines.append(f"- {item}")

    if v.get("fields_to_fill"):
        lines.append("")
        lines.append("DATI DA COMPLETARE")
        for f in v["fields_to_fill"]:
            lines.append(f"- {f}")

    if v.get("short_ready_text"):
        lines.append("")
        lines.append("VERBALE SINTETICO PRONTO")
        lines.append(v["short_ready_text"])

    if procedural_flags.get("verbale_additions") or procedural_flags.get("segnalazioni"):
        lines.append("")
        lines.append("TESTO OPERATIVO DA INSERIRE / RICHIAMARE NEL VERBALE")
        lines.append(f"- Dicitura prontuario: {v['short_ready_text']}")
        if procedural_flags.get("verbale_additions"):
            for item in procedural_flags["verbale_additions"]:
                lines.append(f"- Ulteriore dicitura: {item}")
        if procedural_flags.get("segnalazioni"):
            for item in procedural_flags["segnalazioni"]:
                lines.append(f"- Comunicazione conseguente: {item}")

    if v.get("notes"):
        lines.append("")
        lines.append("NOTE OPERATIVE")
        for n in v["notes"]:
            lines.append(f"- {n}")

    if concurrent_codes:
        lines.append("")
        lines.append("VIOLAZIONI CONCORRENTI POSSIBILI")
        for code in concurrent_codes:
            if code in VIOLATIONS:
                lines.append("")
                lines.append(format_compact_violation(code))
            else:
                lines.append(f"- {code}")

    if ancillary_findings:
        lines.append("")
        lines.append("ULTERIORI INFRAZIONI / ACCERTAMENTI COMPLEMENTARI")
        for item in ancillary_findings:
            lines.append(f"- {item}")

    if procedural_flags.get("segnalazioni"):
        lines.append("")
        lines.append("COMUNICAZIONI / ULTERIORI PROVVEDIMENTI")
        for item in procedural_flags["segnalazioni"]:
            lines.append(f"- {item}")

    if extra_notes:
        lines.append("")
        lines.append("ULTERIORI VERIFICHE")
        for note in extra_notes:
            lines.append(f"- {note}")

    lines.append(article_shortcuts_from_result(main_code, concurrent_codes))

    lines.append("")
    lines.append("AVVERTENZA")
    lines.append("Verificare sempre normativa vigente, prontuario del comando, disciplina locale e dati concreti del caso.")

    return "\n".join(lines)

def format_partial_assessment(answers, concurrent_codes=None, extra_notes=None, procedural_flags=None, ancillary_findings=None):
    if concurrent_codes is None:
        concurrent_codes = []
    if extra_notes is None:
        extra_notes = []
    if procedural_flags is None:
        procedural_flags = {}
    if ancillary_findings is None:
        ancillary_findings = []

    lines = []
    lines.append("ESITO PARZIALE / DA APPROFONDIRE")
    lines.append("Il bot non ha individuato una violazione principale chiudibile con sufficiente affidabilità solo con i dati attuali.")

    filled = []
    labels = {
        "vehicle_authorized": "Veicolo autorizzato NCC",
        "service_to_third": "Servizio verso terzi / utenza",
        "service_context": "Contesto del servizio",
        "violation_type": "Tipo violazione",
        "recurrence": "Progressione violazioni",
        "kb": "Titolo professionale KB/KA/CQC",
        "public_waiting": "Attesa su area pubblica",
        "taxi_commune": "Comune con servizio taxi",
        "booking": "Prenotazione documentabile",
        "separate_payment": "Pagamento separato",
        "rent_registered": "Iscrizione RENT",
        "foglio_status": "Foglio di servizio",
        "ruolo_conducenti": "Ruolo/albo conducenti",
        "patente_idonea": "Patente idonea",
        "incauto_affidamento": "Incauto affidamento"
    }
    for key, label in labels.items():
        value = answers.get(key)
        if value is not None:
            filled.append(f"- {label}: {value}")

    if filled:
        lines.append("")
        lines.append("DATI GIÀ ACQUISITI")
        lines.extend(filled)

    if concurrent_codes:
        lines.append("")
        lines.append("VIOLAZIONI CONCORRENTI / RICHIAMI GIÀ EMERSI")
        for code in concurrent_codes:
            if code in VIOLATIONS:
                lines.append(format_compact_violation(code))
            else:
                lines.append(f"- {code}")

    if ancillary_findings:
        lines.append("")
        lines.append("ULTERIORI INFRAZIONI / ACCERTAMENTI COMPLEMENTARI")
        for item in ancillary_findings:
            lines.append(f"- {item}")

    if procedural_flags.get("verbale_additions"):
        lines.append("")
        lines.append("DICITURE DA RICHIAMARE NEL VERBALE")
        for item in procedural_flags["verbale_additions"]:
            lines.append(f"- {item}")

    if procedural_flags.get("segnalazioni"):
        lines.append("")
        lines.append("COMUNICAZIONI / SEGNALAZIONI CONSEGUENTI")
        for item in procedural_flags["segnalazioni"]:
            lines.append(f"- {item}")

    if extra_notes:
        lines.append("")
        lines.append("VERIFICHE ANCORA NECESSARIE")
        for note in extra_notes:
            lines.append(f"- {note}")

    lines.append("")
    lines.append("AVVERTENZA")
    lines.append("Completa il quadro con ulteriori risposte oppure verifica su prontuario, normativa vigente e disciplina locale.")
    return "\n".join(lines)

def normalize_plate_value(value):
    if value is None:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(value).upper().strip())


def normalize_header_value(value):
    if value is None:
        return ""
    raw = str(value).strip().lower()
    raw = raw.replace("à", "a").replace("è", "e").replace("é", "e").replace("ì", "i").replace("ò", "o").replace("ù", "u")
    raw = raw.replace("_", " ").replace("-", " ")
    raw = re.sub(r"\s+", " ", raw)
    return raw


def _find_first_matching_column(headers, aliases):
    for idx, header in enumerate(headers):
        if header in aliases:
            return idx
    return None


def _interpret_ncc_status(value):
    normalized = normalize_header_value(value)
    compact = normalized.replace(" ", "")

    yes_values = {
        "si", "s", "yes", "y", "true", "1", "ncc", "adibito",
        "abilitato", "autorizzato", "attivo", "presente", "iscritto"
    }
    no_values = {
        "no", "n", "false", "0", "non ncc", "non adibito", "non abilitato",
        "non autorizzato", "revocato", "sospeso", "cessato", "assente"
    }

    if compact in yes_values or normalized in yes_values:
        return "si"
    if compact in no_values or normalized in no_values:
        return "no"
    if "non" in normalized and any(x in normalized for x in ["ncc", "adibito", "autorizzato", "abilitato"]):
        return "no"
    if any(x in normalized for x in ["ncc", "adibito", "autorizzato", "abilitato", "attivo"]):
        return "si"
    return None


def _plate_registry_header_row_index(sheet, max_scan_rows=10):
    target_headers = {
        "targa", "modello", "colore", "destinazione uso veicoli", "uso veicolo",
        "licenza autoveicolo", "intestatario", "residenza intestatario"
    }
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
        normalized = {normalize_header_value(cell) for cell in row if cell not in (None, "")}
        if "targa" in normalized and len(normalized.intersection(target_headers)) >= 3:
            return row_index
    return 1



def _is_persona_fisica_owner(owner_name):
    if owner_name in (None, ""):
        return False
    normalized = normalize_header_value(owner_name).upper()
    company_markers = {
        "SRL", "S R L", "SPA", "S P A", "SAS", "S A S", "SNC", "S N C",
        "SAPA", "S A P A", "COOP", "COOPERATIVA", "SOC", "SOCIETA", "DITTA",
        "IMPRESA", "CONSORZIO", "ASSOCIAZIONE", "FONDAZIONE", "GROUP", "SERVICE",
        "SERVICES", "VIAGGI", "TRAVEL", "TRASPORTI"
    }
    return not any(marker in normalized for marker in company_markers)



def lookup_plate_in_registry(plate_text):
    plate = normalize_plate_value(plate_text)
    if not plate:
        return {"ok": False, "message": "Inserisci una targa valida."}

    if not os.path.exists(TARGHE_FILE_PATH):
        return {
            "ok": False,
            "message": (
                f"Archivio targhe non trovato: {TARGHE_FILE_PATH}."
                "Carica il file Excel nel repository e verifica il percorso in TARGHE_FILE_PATH."
            )
        }

    try:
        from openpyxl import load_workbook
    except Exception as e:
        return {
            "ok": False,
            "message": f"Libreria openpyxl non disponibile sul server: {e}"
        }

    try:
        workbook = load_workbook(TARGHE_FILE_PATH, data_only=True, read_only=True)
        target_sheet_name = TARGHE_SHEET_NAME if TARGHE_SHEET_NAME in workbook.sheetnames else None
        if not target_sheet_name:
            for candidate in workbook.sheetnames:
                if str(candidate).strip().lower() == 'ncc':
                    target_sheet_name = candidate
                    break
        sheet = workbook[target_sheet_name] if target_sheet_name else workbook[workbook.sheetnames[0]]

        header_row_index = _plate_registry_header_row_index(sheet)
        rows = sheet.iter_rows(min_row=header_row_index, values_only=True)
        headers_raw = next(rows, None)
        if not headers_raw:
            return {"ok": False, "message": "Il file Excel targhe è vuoto."}

        headers = [normalize_header_value(h) for h in headers_raw]

        targa_idx = _find_first_matching_column(headers, {"targa", "plate", "telaio/targa", "veicolo", "mezzo"})
        uso_idx = _find_first_matching_column(headers, {"uso veicolo", "uso", "uso del veicolo"})
        intestatario_idx = _find_first_matching_column(headers, {"intestatario", "proprietario", "ragione sociale", "titolare"})
        residenza_idx = _find_first_matching_column(headers, {"residenza intestatario", "residenza", "indirizzo intestatario", "comune intestatario"})
        modello_idx = _find_first_matching_column(headers, {"modello", "veicolo modello", "marca modello"})
        destinazione_idx = _find_first_matching_column(headers, {"destinazione uso veicoli", "destinazione uso", "destinazione"})
        licenza_idx = _find_first_matching_column(headers, {"licenza autoveicolo", "licenza", "autorizzazione", "licenza ncc"})
        note_idx = _find_first_matching_column(headers, {"note", "annotazioni", "osservazioni"})

        if targa_idx is None:
            return {
                "ok": False,
                "message": "Nel file Excel manca una colonna riconoscibile per la targa (es. 'targa')."
            }

        found_row = None
        for row in rows:
            cell_value = row[targa_idx] if targa_idx < len(row) else None
            if normalize_plate_value(cell_value) == plate:
                found_row = row
                break

        if not found_row:
            return {
                "ok": True,
                "found": False,
                "plate": plate,
                "message": f"Il mezzo con targa {plate} non è stato censito."
            }

        def get_value(idx):
            if idx is None or idx >= len(found_row):
                return ""
            value = found_row[idx]
            if value is None:
                return ""
            return str(value).strip()

        uso = get_value(uso_idx).upper()
        intestatario = get_value(intestatario_idx)
        residenza = get_value(residenza_idx)
        modello = get_value(modello_idx)
        destinazione = get_value(destinazione_idx)
        licenza = get_value(licenza_idx)
        note = get_value(note_idx)
        owner_is_person = _is_persona_fisica_owner(intestatario)
        sanctionable = uso == 'PROPRIO' and owner_is_person

        lines = [
            f"Targa: {plate}",
            "",
            "Mezzo presente nel censimento.",
        ]
        if modello:
            lines.append(f"Modello: {modello}")
        if destinazione:
            lines.append(f"Destinazione uso veicolo: {destinazione}")
        if uso:
            lines.append(f"Uso veicolo: {uso}")
        if intestatario:
            lines.append(f"Intestatario: {intestatario}")
        if residenza:
            lines.append(f"Residenza proprietario: {residenza}")
        if licenza:
            lines.append(f"Licenza/autorizzazione: {licenza}")

        lines.append("")
        if sanctionable:
            lines.append("ESITO OPERATIVO: mezzo già sanzionabile.")
            lines.append("Motivo: veicolo censito, uso proprio e intestatario persona fisica.")
        else:
            lines.append("ESITO OPERATIVO: mezzo censito.")
            lines.append("Valutare comunque licenza, foglio di servizio e modalità del trasporto concreto.")

        if note:
            lines.extend(["", f"Note archivio: {note}"])

        return {
            "ok": True,
            "found": True,
            "plate": plate,
            "usage": uso,
            "owner": intestatario,
            "owner_residence": residenza,
            "sanctionable": sanctionable,
            "message": "\n".join(lines)
        }
    except Exception as e:
        return {"ok": False, "message": f"Errore lettura archivio targhe: {e}"}


def begin_plate_lookup_flow(chat_id):
    user_states[chat_id] = {
        "mode": "plate_lookup"
    }
    save_user_states()


def process_plate_lookup(chat_id, text):
    result = lookup_plate_in_registry(text)
    clear_case(chat_id)
    return result


def format_norme_from_db():
    lines = ["RIFERIMENTI NORMATIVI NCC\n"]
    for item in NCC_DB["norme"].values():
        lines.append(f"- {item['norma']}")
        lines.append(f"  Tema: {item['tema']}")
        lines.append(f"  Uso operativo: {item['uso_operativo']}")
        lines.append("")
    return "\n".join(lines).strip()

def format_documenti_from_db():
    lines = ["DOCUMENTI / ELEMENTI DA CONTROLLARE\n"]
    for item in NCC_DB["documenti_controllo"]:
        lines.append(f"- {item}")
    return "\n".join(lines)

def format_checklist_from_db():
    lines = ["CHECKLIST OPERATIVA NCC\n"]
    for i, item in enumerate(NCC_DB["checklist_operativa"], start=1):
        lines.append(f"{i}. {item}")
    return "\n".join(lines)



def begin_port_common_case(chat_id, case_key):
    user_states[chat_id] = {
        "mode": "porto_common_followup",
        "porto_case_key": case_key,
        "answers": {},
        "pending_question": None,
        "questions_asked": []
    }

    state = user_states[chat_id]

    if case_key == "uso_proprio_kb":
        state["answers"].update({
            "vehicle_authorized": "no",
            "service_to_third": "si",
            "kb": "si"
        })
        q = {
            "key": "recurrence_triennio",
            "text": "È la prima violazione oppure la seconda nel triennio?"
        }

    elif case_key == "abusivo_totale":
        state["answers"].update({
            "vehicle_authorized": "no",
            "service_to_third": "si"
        })
        q = {
            "key": "recurrence_triennio",
            "text": "È la prima violazione oppure la seconda nel triennio?"
        }

    elif case_key == "procacciamento":
        state["answers"].update({
            "vehicle_authorized": "si",
            "service_to_third": "si",
            "booking": "no",
            "public_waiting": "si",
            "violation_type": "art3_11"
        })
        q = {
            "key": "recurrence",
            "text": "Indica la progressione della violazione nel quinquennio."
        }

    else:
        return None

    state["pending_question"] = q
    save_user_states()
    return build_recurrence_prompt(state["answers"], q["key"]), q["key"]


def _finalize_port_common_case(chat_id):
    state = user_states.get(chat_id, {})
    answers = state.get("answers", {})

    main_code, concurrent, notes, procedural_flags, ancillary_findings = decide_violation(answers)

    if main_code:
        payload = build_final_payload(
            main_code,
            concurrent_codes=concurrent,
            extra_notes=notes,
            procedural_flags=procedural_flags,
            ancillary_findings=ancillary_findings
        )
        result = payload.get("quick", "Esito finale non disponibile.")
    elif concurrent:
        result = format_partial_assessment(
            answers,
            concurrent,
            notes,
            procedural_flags,
            ancillary_findings
        )
        payload = None
    else:
        result = (
            "Dagli elementi raccolti non emerge, allo stato, una violazione NCC chiudibile automaticamente. "
            "Verificare normativa vigente, disciplina locale e dati concreti del caso."
        )
        payload = None

    clear_case(chat_id)

    if payload:
        state = user_states.setdefault(chat_id, {})
        state["last_result_payload"] = payload
        state["last_result_main_code"] = main_code
        state["last_result_concurrent"] = concurrent
        state["last_result_flags"] = procedural_flags
        save_user_states()

    return result, payload


def process_port_common_followup(chat_id, text):
    state = user_states.get(chat_id)
    if not state:
        return "Sessione non trovata.", None, None

    pending = state.get("pending_question") or {}
    key = pending.get("key")
    value = parse_answer_for_key(key, text) if key in {"recurrence_triennio", "recurrence", "kb", "patente_idonea", "incauto_affidamento"} else normalize_answer(text)

    if key == "recurrence_triennio":
        if value not in {"first", "second_3y"}:
            return "Risposta non valida. Indica: prima / seconda nel triennio.", None, key

        state["answers"]["recurrence_triennio"] = value

        if state.get("porto_case_key") == "abusivo_totale":
            state["pending_question"] = {
                "key": "kb",
                "text": "Il conducente ha KB/KA/CQC richiesto?"
            }
            save_user_states()
            return build_article_verification_prompt(state["answers"], "kb", state["pending_question"]["text"]), None, "kb"

        save_user_states()
        result, payload = _finalize_port_common_case(chat_id)
        return result, payload, None

    if key == "recurrence":
        if value not in {"first", "2_5y", "3_5y", "4plus_5y"}:
            return "Risposta non valida. Indica: prima / seconda / terza / quarta o successiva.", None, key
        state["answers"]["recurrence"] = value
        save_user_states()
        result, payload = _finalize_port_common_case(chat_id)
        return result, payload, None

    if key == "kb":
        if value not in {"si", "no"}:
            return "Risposta non valida. Indica: si / no.", None, key

        state["answers"]["kb"] = value
        state["pending_question"] = {
            "key": "patente_idonea",
            "text": "Il conducente ha patente idonea?"
        }
        save_user_states()
        return build_article_verification_prompt(state["answers"], "patente_idonea", state["pending_question"]["text"]), None, "patente_idonea"

    if key == "patente_idonea":
        if value not in {"si", "no"}:
            return "Risposta non valida. Indica: si / no.", None, key

        state["answers"]["patente_idonea"] = value

        if state["answers"].get("kb") == "no" or value == "no":
            state["pending_question"] = {
                "key": "incauto_affidamento",
                "text": "Il veicolo è stato affidato dal titolare o da altro responsabile a soggetto privo dei titoli richiesti?"
            }
            save_user_states()
            return build_article_verification_prompt(state["answers"], "incauto_affidamento", state["pending_question"]["text"]), None, "incauto_affidamento"

        save_user_states()
        result, payload = _finalize_port_common_case(chat_id)
        return result, payload, None

    if key == "incauto_affidamento":
        if value not in {"si", "no"}:
            return "Risposta non valida. Indica: si / no.", None, key

        state["answers"]["incauto_affidamento"] = value
        save_user_states()
        result, payload = _finalize_port_common_case(chat_id)
        return result, payload, None

    return "Nessuna domanda attiva.", None, None

def format_verbale_template():
    return (
        "SCHEMA RAPIDO VERBALE / ANNOTAZIONE\n\n"
        "1. Descrizione sintetica del fatto accertato.\n"
        "2. Norma principale contestata.\n"
        "3. Eventuali violazioni concorrenti.\n"
        "4. Dicitura integrativa su RENT / ruolo conducenti / segnalazione Prefettura / ente rilasciante.\n"
        "5. Dati da completare: prenotazione, foglio di servizio, passeggeri, titolo autorizzativo, UMC competente.\n\n"
        "Formula utile: 'Si accerta inoltre ...; della circostanza si dà atto nel presente verbale e si procede alla segnalazione alla Prefettura e/o all'ente rilasciante per quanto di competenza.'"
    )

def match_case_from_text(text):
    t = text.lower()

    # casi tipici di primo orientamento
    if any(x in t for x in ["veicolo privato", "auto privata", "mezzo privato", "senza autorizzazione ncc", "abusivo"]):
        return "A_abusivo_totale_mezzo_proprio"

    if any(x in t for x in ["ncc", "prenotazione", "foglio di servizio", "rimessa", "staziona", "sosta"]):
        return "B_titolare_ncc_procaccia_clienti"

    if any(x in t for x in ["agenzia", "impresa", "mezzo aziendale", "dipendente"]):
        return "C_impresa_agenzia_mezzo_aziendale"

    if any(x in t for x in ["navetta", "hotel", "albergo", "parcheggio", "parking", "struttura ricettiva"]):
        return "D_navetta_parcheggio_hotel"

    if any(x in t for x in ["senza kb", "senza cqc", "manca kb", "manca cqc", "senza patente", "patente non idonea"]):
        return "E_veicolo_ncc_conducente_non_legittimato"

    if any(x in t for x in ["con kb", "titolo presente", "conducente titolato"]) and any(x in t for x in ["veicolo privato", "non ncc"]):
        return "F_veicolo_non_ncc_ma_conducente_titolato"

    return None

def format_case_hint(case_key):
    case_data = NCC_DB["casi_tipici"][case_key]
    lines = []
    lines.append("INQUADRAMENTO OPERATIVO PRELIMINARE")
    lines.append(case_data["descrizione"])
    lines.append("")
    lines.append("POSSIBILI ESITI")
    for code in case_data["possibili_esiti"]:
        if code in VIOLATIONS:
            lines.append(f"- {code} | {VIOLATIONS[code]['article']} | {VIOLATIONS[code]['title']}")
        elif code in NCC_DB["norme"]:
            lines.append(f"- {code} | {NCC_DB['norme'][code]['norma']} | {NCC_DB['norme'][code]['tema']}")
        else:
            lines.append(f"- {code}")
    lines.append("")
    lines.append("NOTE")
    for n in case_data["note"]:
        lines.append(f"- {n}")
    return "\n".join(lines)

ARTICOLI_DB = {
    "art85": {
        "titolo": "CdS art. 85",
        "testo": (
            "Art. 85 CdS – Servizio di noleggio con conducente.\n\n"
            "Uso operativo nel bot:\n"
            "- comma 4: veicolo non adibito / non autorizzato a NCC\n"
            "- comma 4-bis: violazioni degli artt. 3 e 11 della L. 21/1992\n"
            "- comma 4-ter: altre prescrizioni dell’autorizzazione NCC\n\n"
            "Per il dettaglio sanzionatorio il bot usa le voci:\n"
            "085-02, 085-04, 085-05, 085-06, 085-07, 085-08, 085-09."
        ),
        "link": "https://www.normattiva.it/"
    },
    "art116": {
        "titolo": "CdS art. 116",
        "testo": (
            "Art. 116 CdS – Requisiti per la guida dei veicoli.\n\n"
            "Uso operativo nel bot:\n"
            "- comma 14: incauto affidamento\n"
            "- commi 15 e 17: guida senza patente / recidiva / reiterazione\n"
            "- comma 15-bis: patente diversa in casi tipizzati\n"
            "- commi 16 e 18: guida senza CAP / KB / KA / CQC\n\n"
            "Per il dettaglio operativo il bot usa le voci:\n"
            "116-01, 116-02, 116-03, 116-04, 116-05, 116-06."
        ),
        "link": "https://www.normattiva.it/"
    },
    "art3l21": {
        "titolo": "L. 21/1992 art. 3",
        "testo": (
            "Legge 21/1992 – art. 3.\n\n"
            "Uso operativo nel bot:\n"
            "- la richiesta del servizio deve essere avanzata presso la sede o la rimessa, "
            "anche mediante strumenti tecnologici;\n"
            "- lo stazionamento dei mezzi deve avvenire all’interno delle rimesse;\n"
            "- sede operativa e almeno una rimessa devono essere nel territorio consentito.\n\n"
            "Se il mezzo è regolarmente NCC ma viola queste prescrizioni, il bot orienta verso il ramo 085-05 / 085-06 / 085-07 / 085-08."
        ),
        "link": "https://www.normattiva.it/"
    },
    "art11l21": {
        "titolo": "L. 21/1992 art. 11",
        "testo": (
            "Legge 21/1992 – art. 11.\n\n"
            "Uso operativo nel bot:\n"
            "- divieto di sosta in posteggio di stazionamento su suolo pubblico nei comuni dove è esercitato il servizio taxi;\n"
            "- nei casi consentiti, verifica dell’area di sosta autorizzata;\n"
            "- controllo su prenotazione / foglio di servizio / modalità operative del servizio.\n\n"
            "Se il mezzo è NCC e la violazione ricade su queste prescrizioni, il bot orienta verso il ramo 085-05 / 085-06 / 085-07 / 085-08."
        ),
        "link": "https://www.normattiva.it/"
 
    },
    "art180": {
        "titolo": "CdS art. 180",
        "testo": (
            "Art. 180 CdS – Possesso ed esibizione dei documenti di circolazione e di guida.\n\n"
            "Uso operativo nel bot:\n"
            "- si applica quando il documento esiste ma non viene esibito all'atto del controllo;\n"
            "- va distinto dalle violazioni sostanziali, cioe quando il titolo manca, e revocato, sospeso o non idoneo;\n"
            "- nel bot richiama soprattutto le ipotesi documentali per patente, autorizzazione NCC, foglio di servizio, carta di circolazione e assicurazione."
        ),
        "link": "https://www.normattiva.it/"
    },
    "art126": {
        "titolo": "CdS art. 126",
        "testo": (
            "Art. 126 CdS – Durata e conferma della validita della patente di guida.\n\n"
            "Uso operativo nel bot:\n"
            "- rileva quando patente, CAP, KB o CQC risultano scaduti;\n"
            "- va distinto dalla mancanza del titolo e dalla mera mancata esibizione;\n"
            "- nel bot e usato come richiamo operativo per i titoli scaduti."
        ),
        "link": "https://www.normattiva.it/"
   }
}

# =========================
# ANALISI TESTO LIBERO
# =========================

def _contains_any(text, needles):
    return any(x in text for x in needles)


def _append_unique(lst, value):
    if value and value not in lst:
        lst.append(value)


def detect_from_text(text):
    t = _normalize_free_answer(text)

    data = {
        "vehicle_authorized": None,
        "service_to_third": None,
        "service_context": None,
        "violation_type": None,
        "recurrence": None,
        "kb": None,
        "public_waiting": None,
        "taxi_commune": None,
        "booking": None,
        "separate_payment": None,
        "rent_registered": None,
        "foglio_status": None,
        "ruolo_conducenti": None,
        "patente_idonea": None,
        "incauto_affidamento": None,
        "ente_rilasciante_known": None
    }

    if _contains_any(t, [
        "veicolo privato", "auto privata", "macchina privata", "mezzo privato",
        "senza autorizzazione ncc", "non autorizzato ncc", "abusivo",
        "non ncc", "veicolo non ncc", "auto non ncc", "veicolo ad uso proprio",
        "vettura privata", "targa civile non ncc"
    ]):
        data["vehicle_authorized"] = "no"

    if _contains_any(t, [
        "veicolo ncc", "autorizzato ncc", "con autorizzazione ncc",
        "mezzo ncc", "auto ncc", "ncc regolare", "licenza ncc",
        "autorizzazione comunale", "autovettura ncc"
    ]):
        data["vehicle_authorized"] = "si"

    if _contains_any(t, [
        "clienti", "passeggeri", "turisti", "utenza", "trasporta persone",
        "porta persone", "accompagna clienti", "accompagna turisti",
        "prende clienti", "fa la corsa", "servizio a pagamento", "trasporto terzi",
        "procaccia clienti", "utenza indifferenziata", "crocieristi", "arrivi",
        "meet and greet", "pickup", "chiama clienti", "stava trasportando",
        "stava caricando passeggeri", "stava scaricando passeggeri"
    ]):
        data["service_to_third"] = "si"

    if _contains_any(t, [
        "trasporto interno", "servizio per clienti propri", "solo clienti hotel",
        "solo clienti del parcheggio", "attivita accessoria", "navetta interna",
        "navetta dell'hotel", "servizio interno", "servizio di cortesia"
    ]):
        data["service_to_third"] = "dubbio"

    if _contains_any(t, ["hotel", "albergo", "parcheggio", "parking", "struttura ricettiva", "navetta", "shuttle", "resort", "b&b", "bed and breakfast"]):
        data["service_context"] = "b"
    elif _contains_any(t, ["taxi", "ncc", "autobus autorizzato", "bus autorizzato"]):
        data["service_context"] = "a"
    else:
        data["service_context"] = "c"

    if _contains_any(t, [
        "pagamento", "a pagamento", "si fa pagare", "corrispettivo",
        "prezzo", "tariffa", "contanti", "pagano il trasporto", "bonifico", "pos"
    ]) or re.search(r"\d+[\.,]?\d*\s*euro", t):
        data["separate_payment"] = "si"

    if _contains_any(t, ["gratuito", "senza corrispettivo", "compreso nel servizio", "servizio incluso", "cortesia"]):
        data["separate_payment"] = "no"

    if _contains_any(t, ["senza prenotazione", "no prenotazione", "manca prenotazione", "privo di prenotazione"]):
        data["booking"] = "no"
    if _contains_any(t, ["con prenotazione", "prenotazione presente", "prenotazione documentabile", "contratto presente"]):
        data["booking"] = "si"

    if _contains_any(t, ["foglio di servizio assente", "senza foglio di servizio", "manca foglio di servizio", "foglio non compilato"]):
        data["foglio_status"] = "assente"
        data["booking"] = data["booking"] or "no"
        data["violation_type"] = "art3_11"
    elif _contains_any(t, ["foglio di servizio irregolare", "foglio incompleto", "foglio compilato male", "foglio irregolare"]):
        data["foglio_status"] = "irregolare"
        data["violation_type"] = "art3_11"
    elif _contains_any(t, ["foglio di servizio non esibito", "non esibisce il foglio", "rifiuta di esibire il foglio"]):
        data["foglio_status"] = "non_esibito"
        if data["booking"] is None:
            data["booking"] = "si"
    elif _contains_any(t, ["foglio di servizio presente", "esibisce il foglio", "foglio regolare"]):
        data["foglio_status"] = "presente"
        if data["booking"] is None:
            data["booking"] = "si"

    if _contains_any(t, [
        "staziona", "sosta su area pubblica", "attesa clienti", "fuori rimessa", "in attesa al porto", "in attesa in aeroporto", "in attesa in stazione",
        "in attesa al terminal", "in attesa in strada", "senza prenotazione", "procaccia clienti", "utenza indifferenziata"
    ]):
        data["violation_type"] = "art3_11"

    if _contains_any(t, ["ztl", "regolamento comunale", "prescrizione autorizzativa", "altra prescrizione", "condizioni autorizzazione"]):
        data["violation_type"] = "other_auth"

    if data["violation_type"] is None:
        data["violation_type"] = "none"

    if _contains_any(t, ["staziona", "sosta", "in attesa", "fermo in attesa", "attesa clienti", "fuori terminal", "fuori porto", "davanti al terminal", "su area pubblica", "procaccia clienti", "utenza indifferenziata", "davanti hotel", "davanti aeroporto", "davanti stazione"]):
        data["public_waiting"] = "si"

    if _contains_any(t, ["porto di civitavecchia", "roma", "fiumicino", "milano", "napoli", "stazione termini", "aeroporto", "porto", "terminal crociere", "molo", "stazione"]):
        data["taxi_commune"] = "si"

    if _contains_any(t, [
        "senza kb", "manca kb", "privo di kb", "kb assente", "licenza kb assente", "senza licenza kb",
        "senza titolo kb", "titolo kb assente", "senza ka", "manca ka", "senza cqc", "manca cqc",
        "privo di cqc", "privo di ka", "cap kb mancante"
    ]):
        data["kb"] = "no"
    if _contains_any(t, ["con kb", "kb presente", "cqc presente", "titolo presente", "cap kb presente", "titolo kb presente", "licenza kb presente"]):
        data["kb"] = "si"

    if _contains_any(t, ["patente sospesa", "senza patente", "patente revocata", "patente non idonea", "patente scaduta"]):
        data["patente_idonea"] = "no"
    if _contains_any(t, ["patente regolare", "patente valida", "patente idonea"]):
        data["patente_idonea"] = "si"

    if _contains_any(t, ["non iscritto al rent", "manca rent", "assenza rent", "rent non registrato", "rent assente", "non risulta rent"]):
        data["rent_registered"] = "no"
    if _contains_any(t, ["rent presente", "iscritto al rent", "registrato rent"]):
        data["rent_registered"] = "si"

    if _contains_any(t, ["non iscritto al ruolo", "manca iscrizione al ruolo", "ruolo conducenti assente", "non iscritto albo conducenti", "albo conducenti assente", "ruolo assente"]):
        data["ruolo_conducenti"] = "no"
    if _contains_any(t, ["iscritto al ruolo", "ruolo conducenti presente", "albo conducenti presente"]):
        data["ruolo_conducenti"] = "si"

    if _contains_any(t, ["incauto affidamento", "affidato dal titolare", "veicolo affidato", "dato in uso dal titolare"]):
        data["incauto_affidamento"] = "si"

    if _contains_any(t, ["comune rilasciante noto", "ente rilasciante noto", "autorizzazione del comune di"]):
        data["ente_rilasciante_known"] = "si"

    if _contains_any(t, ["seconda nel triennio", "2a nel triennio", "recidiva triennio"]):
        data["recurrence"] = "second_3y"
    elif _contains_any(t, ["seconda nel quinquennio", "2a nel quinquennio"]):
        data["recurrence"] = "2_5y"
    elif _contains_any(t, ["terza nel quinquennio", "3a nel quinquennio"]):
        data["recurrence"] = "3_5y"
    elif _contains_any(t, ["quarta nel quinquennio", "quarta o successiva", "4a nel quinquennio"]):
        data["recurrence"] = "4plus_5y"
    elif _contains_any(t, ["prima violazione", "1a violazione", "prima volta"]):
        data["recurrence"] = "first"

    return data


def decide_violation(answers):
    vehicle_authorized = answers.get("vehicle_authorized")
    service_to_third = answers.get("service_to_third")
    service_context = answers.get("service_context")
    violation_type = answers.get("violation_type")
    recurrence = answers.get("recurrence")
    recurrence_triennio = answers.get("recurrence_triennio") or recurrence
    kb = answers.get("kb")
    public_waiting = answers.get("public_waiting")
    taxi_commune = answers.get("taxi_commune")
    booking = answers.get("booking")
    separate_payment = answers.get("separate_payment")
    rent_registered = answers.get("rent_registered")
    foglio_status = answers.get("foglio_status")
    ruolo_conducenti = answers.get("ruolo_conducenti")
    patente_idonea = answers.get("patente_idonea")
    incauto_affidamento = answers.get("incauto_affidamento")
    owner_type = answers.get("owner_type")
    circulation_use = answers.get("circulation_use")
    trip_nature = answers.get("trip_nature")

    concurrent = []
    notes = []
    ancillary_findings = []
    procedural_flags = {"segnalazioni": [], "verbale_additions": []}

    def add_signal(text):
        _append_unique(procedural_flags["segnalazioni"], text)

    def add_verbal(text):
        _append_unique(procedural_flags["verbale_additions"], text)

    if kb == "no":
        _append_unique(concurrent, "116-06")
        add_verbal("Accertato che il conducente era privo del prescritto titolo professionale KB/KA/CQC ove richiesto.")

    if patente_idonea == "no":
        _append_unique(concurrent, "116-02")
        ancillary_findings.append("Conducente privo di patente/titolo di guida idoneo: verificare se ricorrono recidiva biennale, reiterazione o fattispecie 15-bis.")
        add_verbal("Accertato che il conducente era privo di patente valida o idonea al veicolo/servizio controllato.")

    if incauto_affidamento == "si" and (kb == "no" or patente_idonea == "no"):
        _append_unique(concurrent, "116-01")
        add_verbal("Il veicolo risultava affidato dal titolare o responsabile a soggetto privo dei titoli richiesti; si procede per l'incauto affidamento.")

    if rent_registered == "no":
        add_signal("Comunicazione alla Prefettura per mancata iscrizione RENT, da richiamare nel verbale insieme alle eventuali violazioni contestate.")
        add_signal("Comunicazione al Comune/ente rilasciante per le valutazioni sulla regolarità del titolo NCC e sugli eventuali provvedimenti di competenza.")
        add_verbal("Si accerta la mancata iscrizione al RENT; della circostanza si dà atto nel presente verbale e si procede alla comunicazione alla Prefettura e al Comune/ente rilasciante.")
        ancillary_findings.append("Mancata iscrizione RENT: anomalia amministrativa da segnalazione, da cumulare nel verbale con le altre violazioni contestate.")

    if ruolo_conducenti == "no":
        add_signal("Comunicazione alla Prefettura per conducente privo di iscrizione al ruolo/albo conducenti.")
        add_signal("Comunicazione al Comune/ente rilasciante per verifica dei requisiti soggettivi del conducente e del titolo NCC.")
        add_verbal("Il conducente risultava privo di iscrizione al ruolo/albo conducenti; della circostanza si dà atto nel verbale e si procede alle comunicazioni di competenza.")
        ancillary_findings.append("Conducente non iscritto al ruolo/albo: requisito soggettivo mancante da evidenziare nel verbale e da segnalare.")

    if foglio_status == "assente":
        violation_type = "art3_11"
        booking = "no"
        add_verbal("Foglio di servizio assente o non compilato al momento del controllo.")
        ancillary_findings.append("Foglio di servizio assente/non compilato: integra violazione sostanziale delle modalità di esercizio del servizio.")
    elif foglio_status == "irregolare":
        violation_type = "art3_11"
        add_verbal("Foglio di servizio irregolare/incompleto rispetto al servizio in corso.")
        ancillary_findings.append("Foglio di servizio irregolare/incompleto: trattare come violazione sostanziale delle modalità di esercizio del servizio.")
    elif foglio_status == "non_esibito":
        ancillary_findings.append("Foglio di servizio esistente ma non esibito: valutare separatamente la mancata esibizione documentale ex art. 180 CdS.")
        add_verbal("Il conducente non esibiva nell'immediatezza il foglio di servizio/codice identificativo del servizio; valutare autonoma contestazione documentale.")
        _append_unique(concurrent, "180-01")

    if public_waiting == "si" and booking == "no" and vehicle_authorized == "si":
        violation_type = "art3_11"
        add_verbal("Il veicolo stazionava o sostava su area pubblica in attesa di utenza senza prenotazione documentabile.")

    if circulation_use == "uso_proprio" and service_to_third == "si":
        if owner_type == "agenzia_viaggi" and trip_nature == "agenzia_pacchetto":
            notes.extend([
                "Il mezzo risulta in uso proprio ma il servizio appare collegato a pacchetto/escursione organizzata da agenzia.",
                "Acquisire voucher, contratto, programma del viaggio, corrispettivo complessivo e prova che il trasporto sia accessorio al servizio turistico.",
                "In questo scenario il caso non va chiuso automaticamente come NCC abusivo senza prima verifica documentale completa."
            ])
            add_verbal("Nel corso del controllo il trasporto dichiarato risultava inserito in iniziativa organizzata da agenzia di viaggi; si acquisiva/documentava la relativa documentazione commerciale al fine di verificare se il trasporto costituisse prestazione accessoria a pacchetto turistico ovvero autonomo servizio di trasporto persone.")
            return None, concurrent, notes, procedural_flags, ancillary_findings
        if trip_nature == "navetta_accessoria":
            notes.extend([
                "Il mezzo risulta in uso proprio e il servizio appare come navetta/accessorio.",
                "Verificare che sia riservato a clienti propri, senza corrispettivo separato per il trasporto e senza apertura a utenza indifferenziata."
            ])
            add_verbal("Il trasporto appariva dichiarato come servizio accessorio/navetta: acquisire elementi su prenotazione, platea dei destinatari e corrispettivo del trasporto prima di contestare la violazione principale.")
            return None, concurrent, notes, procedural_flags, ancillary_findings
        add_verbal("Accertato servizio di trasporto persone verso corrispettivo con veicolo risultante dal documento di circolazione adibito a uso proprio.")
        add_signal("Comunicazione al Prefetto del luogo della violazione nei casi di art. 85 comma 4 CdS.")
        ancillary_findings.append("Uso proprio sul documento di circolazione incompatibile, allo stato, con servizio NCC/trasporto verso terzi a pagamento: verificare e verbalizzare dettagli del DU/libretto e del rapporto con i passeggeri.")
        vehicle_authorized = "no"

    if service_to_third == "no" and vehicle_authorized == "si" and violation_type in {None, "none"} and not concurrent:
        return None, concurrent, [
            "Dagli elementi forniti non emerge al momento una violazione NCC tipica già chiudibile con art. 85.",
            "Verificare comunque documenti, prenotazione, foglio di servizio e prescrizioni locali."
        ], procedural_flags, ancillary_findings

    if service_context == "b" and separate_payment == "no":
        return None, concurrent, [
            "Il caso può rientrare in navetta / trasporto accessorio collegato ad attività propria.",
            "Verificare che il servizio sia riservato a clienti propri della struttura o attività.",
            "Verificare che non vi sia corrispettivo separato specifico per il trasporto.",
            "Verificare che non si tratti in concreto di servizio aperto a utenza indifferenziata."
        ], procedural_flags, ancillary_findings

    if vehicle_authorized == "no" and service_to_third == "si":
        add_signal("Trasmettere il verbale al Prefetto del luogo della violazione nei casi previsti dall'art. 85 comma 4 CdS.")
        add_verbal("Accertato servizio di noleggio con conducente svolto con veicolo non adibito/autorizzato a tale uso.")
        if recurrence_triennio == "second_3y":
            add_signal("Comunicazione al Prefetto entro 5 giorni per i presupposti della revoca patente, trattandosi di seconda violazione utile nel triennio.")
            return "085-04", concurrent, notes, procedural_flags, ancillary_findings
        return "085-02", concurrent, notes, procedural_flags, ancillary_findings

    if vehicle_authorized == "si" and violation_type == "art3_11":
        add_verbal("Condotta ricondotta a violazione delle modalità di esercizio del servizio NCC di cui agli artt. 3 e/o 11 L. 21/1992.")
        add_signal("Comunicazione all'UMC competente per il ritiro/trasmissione del documento di circolazione nei casi di sospensione previsti dall'art. 85 comma 4-bis CdS.")
        add_signal("Valutare comunicazione al Comune/ente rilasciante quando la violazione incide sulle modalità di esercizio o sulla regolarità del titolo NCC.")
        if recurrence == "2_5y":
            return "085-06", concurrent, notes, procedural_flags, ancillary_findings
        elif recurrence == "3_5y":
            return "085-07", concurrent, notes, procedural_flags, ancillary_findings
        elif recurrence == "4plus_5y":
            return "085-08", concurrent, notes, procedural_flags, ancillary_findings
        else:
            return "085-05", concurrent, notes, procedural_flags, ancillary_findings

    if vehicle_authorized == "si" and violation_type == "other_auth":
        add_verbal("Accertata violazione di prescrizioni o condizioni ulteriori contenute nell'autorizzazione NCC o nella disciplina locale applicabile.")
        add_signal("Valutare comunicazione al Comune/ente rilasciante per eventuali provvedimenti sul titolo autorizzativo.")
        return "085-09", concurrent, notes, procedural_flags, ancillary_findings

    if vehicle_authorized == "si" and public_waiting == "si" and taxi_commune == "si" and booking == "no":
        notes.extend([
            "Possibile violazione art. 11 L. 21/1992 per stazionamento fuori rimessa.",
            "Verificare se il veicolo era in attesa di utenza indifferenziata o già in servizio su prenotazione."
        ])
        add_verbal("Il veicolo NCC sostava/stazionava in comune con servizio taxi attivo, fuori rimessa e senza prenotazione documentabile.")
        add_signal("Comunicazione all'UMC competente per gli adempimenti sul documento di circolazione in caso di contestazione ex art. 85 comma 4-bis.")
        return "085-05", concurrent, notes, procedural_flags, ancillary_findings

    return None, concurrent, [
        "Caso non chiudibile automaticamente.",
        "Servono ulteriori elementi su autorizzazione, prenotazione, natura del servizio e progressione."
    ], procedural_flags, ancillary_findings


def missing_questions(answers):
    questions = []

    vehicle_authorized = answers.get("vehicle_authorized")
    service_to_third = answers.get("service_to_third")
    violation_type = answers.get("violation_type")
    foglio_status = answers.get("foglio_status")

    if vehicle_authorized is None:
        return [{
            "key": "vehicle_authorized",
            "text": "Il veicolo era regolarmente autorizzato/adibito a NCC?\nRispondi: si / no"
        }]

    if service_to_third is None:
        return [{
            "key": "service_to_third",
            "text": "Il conducente stava trasportando o mettendosi a disposizione di clienti/passeggeri?\nRispondi: si / no / dubbio"
        }]

    if vehicle_authorized == "no":
        if service_to_third == "si" and answers.get("recurrence") is None:
            return [{
                "key": "recurrence_triennio",
                "text": "Per il ramo art. 85 c.4 si tratta di prima violazione o seconda nel triennio?\nRispondi: first / second_3y"
            }]
        if answers.get("kb") is None:
            questions.append({
                "key": "kb",
                "text": "Il conducente aveva il titolo professionale richiesto (KB / KA / CQC se dovuto)?\nRispondi: si / no"
            })
        if answers.get("patente_idonea") is None:
            questions.append({
                "key": "patente_idonea",
                "text": "La patente del conducente era valida e idonea al veicolo/servizio?\nRispondi: si / no"
            })
        if (answers.get("kb") == "no" or answers.get("patente_idonea") == "no") and answers.get("incauto_affidamento") is None:
            questions.append({
                "key": "incauto_affidamento",
                "text": "Il veicolo è stato affidato dal titolare o da altro responsabile al conducente privo dei titoli richiesti?\nRispondi: si / no"
            })
        return questions

    if foglio_status is None:
        questions.append({
            "key": "foglio_status",
            "text": "Situazione foglio di servizio?\nRispondi con una sola opzione: presente / assente / irregolare / non_esibito"
        })

    if answers.get("rent_registered") is None:
        questions.append({
            "key": "rent_registered",
            "text": "Il vettore/titolo risultava iscritto al RENT?\nRispondi: si / no"
        })

    if answers.get("ruolo_conducenti") is None:
        questions.append({
            "key": "ruolo_conducenti",
            "text": "Il conducente risultava iscritto al ruolo/albo conducenti quando richiesto?\nRispondi: si / no"
        })

    if answers.get("kb") is None:
        questions.append({
            "key": "kb",
            "text": "Il conducente aveva il titolo professionale richiesto (KB / KA / CQC se dovuto)?\nRispondi: si / no"
        })

    if answers.get("patente_idonea") is None:
        questions.append({
            "key": "patente_idonea",
            "text": "La patente del conducente era valida e idonea al veicolo/servizio?\nRispondi: si / no"
        })

    if violation_type in [None, "none"] and foglio_status not in {"assente", "irregolare"}:
        questions.append({
            "key": "violation_type",
            "text": "Il problema riguarda soprattutto:\nart3_11 = prenotazione / stazionamento / foglio di servizio / rimessa\nother_auth = altre prescrizioni dell'autorizzazione\nnone = non chiaro\nRispondi: art3_11 / other_auth / none"
        })

    if (violation_type == "art3_11" or foglio_status in {"assente", "irregolare"}) and answers.get("recurrence") is None:
        questions.append({
            "key": "recurrence",
            "text": "Per il ramo art. 85 c.4-bis questa violazione è:\nfirst = prima\n2_5y = seconda nel quinquennio\n3_5y = terza nel quinquennio\n4plus_5y = quarta o successiva\nRispondi con una di queste opzioni."
        })

    if service_to_third != "no" and answers.get("public_waiting") is None:
        questions.append({
            "key": "public_waiting",
            "text": "Il veicolo era fermo o in attesa su area pubblica?\nRispondi: si / no"
        })

    if service_to_third != "no" and answers.get("taxi_commune") is None:
        questions.append({
            "key": "taxi_commune",
            "text": "Il fatto è avvenuto in un comune dove è attivo il servizio taxi?\nRispondi: si / no"
        })

    if service_to_third != "no" and answers.get("booking") is None:
        questions.append({
            "key": "booking",
            "text": "C'era una prenotazione documentabile o un titolo di corsa verificabile?\nRispondi: si / no"
        })

    if answers.get("separate_payment") is None and service_to_third != "no":
        questions.append({
            "key": "separate_payment",
            "text": "Per il trasporto era previsto un pagamento o corrispettivo separato?\nRispondi: si / no"
        })

    if (answers.get("kb") == "no" or answers.get("patente_idonea") == "no") and answers.get("incauto_affidamento") is None:
        questions.append({
            "key": "incauto_affidamento",
            "text": "Il veicolo è stato affidato dal titolare o da altro responsabile al conducente privo dei titoli richiesti?\nRispondi: si / no"
        })

    return questions


def _normalize_free_answer(text):
    t = (text or "").strip().lower()
    t = t.replace("sì", "si")
    t = re.sub(r"[\.,;:!\?]+$", "", t)
    t = re.sub(r"\s+", " ", t)
    return t


def _extract_yes_no(text):
    t = _normalize_free_answer(text)
    if re.match(r"^(si|yes|y|certo|confermo|esatto|ok)\b", t):
        return "si"
    if re.match(r"^(no|n|negativo|assolutamente no|non presente|manca|assente)\b", t):
        return "no"
    return None


def _extract_choice(text, allowed):
    t = _normalize_free_answer(text)
    if t in allowed:
        return t
    compact = t.replace(" ", "_").replace("-", "_")
    if compact in allowed:
        return compact
    for choice in allowed:
        if re.search(rf"\b{re.escape(choice)}\b", t):
            return choice
    return None


def _extract_recurrence(text):
    t = _normalize_free_answer(text).replace("^", "")
    mapping = {
        "first": ["first", "prima", "prima violazione", "1a", "1", "prima volta", "prima quinq", "prima nel quinquennio"],
        "second_3y": ["second_3y", "seconda nel triennio", "2a nel triennio", "recidiva triennio", "2 triennio"],
        "2_5y": ["2_5y", "seconda nel quinquennio", "2a nel quinquennio", "2 quinquennio", "seconda"],
        "3_5y": ["3_5y", "terza nel quinquennio", "3a nel quinquennio", "3 quinquennio", "terza"],
        "4plus_5y": ["4plus_5y", "quarta o successiva", "quarta nel quinquennio", "4a nel quinquennio", "quarta+", "4plus", "4", "quarta"],
    }
    for code, vals in mapping.items():
        for v in vals:
            if t == v or re.search(rf"\b{re.escape(v)}\b", t):
                return code
    return None

def _extract_foglio_status(text):
    t = _normalize_free_answer(text)
    if _contains_any(t, ["non esibito", "non_esibito", "non esibisce", "rifiuta di esibire"]):
        return "non_esibito"
    if _contains_any(t, ["assente", "manca", "mancante", "senza foglio", "non compilato"]):
        return "assente"
    if _contains_any(t, ["irregolare", "incompleto", "compilato male", "difforme"]):
        return "irregolare"
    if _contains_any(t, ["presente", "regolare", "esibito", "compilato"]):
        return "presente"
    return None


def parse_answer_for_key(key, text):
    t = text.strip().lower()

    yes_no_keys = [
        "vehicle_authorized", "kb", "public_waiting", "taxi_commune", "booking",
        "separate_payment", "rent_registered", "ruolo_conducenti", "patente_idonea",
        "incauto_affidamento", "ente_rilasciante_known"
    ]
    if key in yes_no_keys:
        yn = _extract_yes_no(t)
        if yn is not None:
            return yn

    if key == "service_to_third":
        if re.search(r"(dubbio|forse|non chiaro|incerto)", _normalize_free_answer(t)):
            return "dubbio"
        yn = _extract_yes_no(t)
        if yn is not None:
            return yn
        if _contains_any(t, ["clienti", "passeggeri", "turisti", "utenza", "trasporto terzi"]):
            return "si"

    if key == "service_context":
        choice = _extract_choice(t, {"a", "b", "c"})
        if choice:
            return choice

    if key == "violation_type":
        choice = _extract_choice(t, {"art3_11", "other_auth", "none"})
        if choice:
            return choice
        if _contains_any(t, ["foglio", "prenotazione", "rimessa", "stazionamento", "porto", "terminal", "utenza indifferenziata"]):
            return "art3_11"
        if _contains_any(t, ["prescrizione", "regolamento", "ztl", "autorizzazione"]):
            return "other_auth"
        yn = _extract_yes_no(t)
        if yn == "no":
            return "none"

    if key in {"recurrence", "recurrence_triennio"}:
        rec = _extract_recurrence(t)
        if rec:
            return rec

    if key == "foglio_status":
        fs = _extract_foglio_status(t)
        if fs:
            return fs

    custom_map = {
        "control_patente_status": {"valida", "scaduta", "non_idonea", "non_esibita"},
        "control_kb_status": {"valido", "scaduto", "non_idoneo", "non_esibito", "non_dovuto"},
        "control_autorizzazione_status": {"regolare", "non_esibita", "non_autorizzato"},
        "control_foglio_status": {"regolare", "irregolare", "assente", "non_esibito"},
        "control_rent_status": {"si", "no", "non_verificato"},
        "control_ruolo_status": {"si", "no", "non_verificato"},
        "control_owner_type": {"persona_fisica", "cooperativa_srl", "agenzia_viaggi", "altro"},
        "control_circulation_use": {"uso_terzi_ncc", "uso_proprio", "non_letto"},
        "control_trip_nature": {"ncc_puro", "agenzia_pacchetto", "navetta_accessoria", "dubbio"},
    }
    if key in custom_map:
        choice = _extract_choice(t, custom_map[key])
        if choice:
            return choice

    return None


def merge_detected_answers(state, text):
    detected = detect_from_text(text)
    for key, value in detected.items():
        if value is not None:
            state["answers"][key] = value

def begin_control_flow(chat_id):
    user_states[chat_id] = {
        "mode": "control_docs",
        "answers": {},
        "selected_docs": [],
        "control_queue": [],
        "pending_question": None,
        "control_concurrent": [],
        "control_notes": [],
        "control_flags": {"segnalazioni": [], "verbale_additions": []},
        "preset_name": None,
        "questions_asked": [],
    }
    save_user_states()


def _control_text_from_state(state):
    selected = state.get("selected_docs", [])
    selected_labels = [CONTROL_DOC_LABELS.get(k, k) for k in selected]
    lines = [
        "CONTROLLO DOCUMENTALE NCC - CHECKLIST OPERATIVA",
        "",
        "Premi i documenti/materiali che il conducente ha esibito.",
        "Quando hai finito premi CONFERMA DOCUMENTI.",
        "",
        "Selezionati:"
    ]
    if selected_labels:
        lines.extend([f"- {x}" for x in selected_labels])
    else:
        lines.append("- nessuno")
    lines.append("")
    lines.append("I documenti non selezionati verranno trattati come non esibiti/mancanti. Il bot farà solo le domande strettamente necessarie per distinguere tra mancata esibizione, mancanza sostanziale, documento scaduto/non valido e, se serve, uso proprio/uso terzi o servizio tramite agenzia.")
    return "\n".join(lines)


def build_control_docs_markup(state):
    selected = set(state.get("selected_docs", []))
    markup = types.InlineKeyboardMarkup(row_width=2)
    buttons = []
    for item in CONTROL_DOCS:
        active = item["id"] in selected
        label = f"✅ {item['label']}" if active else item["label"]
        buttons.append(types.InlineKeyboardButton(label, callback_data=f"ctrl_doc_toggle:{item['id']}"))
    for i in range(0, len(buttons), 2):
        markup.row(*buttons[i:i+2])
    markup.row(types.InlineKeyboardButton("CONFERMA DOCUMENTI", callback_data="ctrl_doc_done"))
    markup.row(types.InlineKeyboardButton("ANNULLA", callback_data="ctrl_doc_cancel"))
    return markup


def _append_unique_local(lst, item):
    if item and item not in lst:
        lst.append(item)


def apply_control_defaults_from_selection(state):
    answers = state.setdefault("answers", {})
    selected = set(state.get("selected_docs", []))

    # Default operativo: il controllo /controllo nasce per servizi NCC,
    # salvo successiva emersione di uso proprio/servizio accessorio o titolo assente.
    for k, v in CONTROL_ASSUME_NCC_DEFAULTS.items():
        answers.setdefault(k, v)

    # Memorizza quali documenti sono stati esibiti: questo evita domande ridondanti
    # e consente di chiudere correttamente i rami documentali ex art. 180.
    answers["doc_patente_esibita"] = "si" if "patente" in selected else "no"
    answers["doc_kb_esibito"] = "si" if "kb" in selected else "no"
    answers["doc_autorizzazione_esibita"] = "si" if "autorizzazione" in selected else "no"
    answers["doc_carta_esibita"] = "si" if "carta" in selected else "no"
    answers["doc_assicurazione_esibita"] = "si" if "assicurazione" in selected else "no"
    answers["doc_foglio_esibito"] = "si" if "foglio" in selected else "no"

    # Se carta/assicurazione non sono state selezionate, predisponi subito la nota documentale.
    concurrent = state.setdefault("control_concurrent", [])
    notes = state.setdefault("control_notes", [])
    if "carta" not in selected:
        if "180-01DOC" not in concurrent:
            concurrent.append("180-01DOC")
        note = "Documento di circolazione / DU non esibito all'atto del controllo: se esistente, contestare art. 180 c.1 e c.7 con invito a presentazione."
        if note not in notes:
            notes.append(note)
    if "assicurazione" not in selected:
        if "180-03" not in concurrent:
            concurrent.append("180-03")
        note = "Certificato/documento assicurativo non esibito all'atto del controllo: contestare art. 180 c.1 e c.7 con invito a presentazione, salvo verifica immediata positiva."
        if note not in notes:
            notes.append(note)

def _queue_control_question(state, key, text):
    state.setdefault("control_queue", []).append({"key": key, "text": text})


def build_control_queue(state):
    state["control_queue"] = []
    selected = set(state.get("selected_docs", []))

    if "patente" in selected:
        _queue_control_question(state, "control_patente_status", "Patente esibita: è valida e idonea al veicolo/servizio?\nScegli: valida / scaduta / non_idonea / non_esibita")
    else:
        _queue_control_question(state, "control_patente_status", "Patente non esibita: si tratta di mera mancata esibizione oppure di patente scaduta o non idonea?\nScegli: non_esibita / scaduta / non_idonea")

    if "kb" in selected:
        _queue_control_question(state, "control_kb_status", "KB / KA / CQC esibito: scegli lo stato corretto.\nScegli: valido / scaduto / non_idoneo / non_dovuto")
    else:
        _queue_control_question(state, "control_kb_status", "KB / KA / CQC non esibito: scegli lo stato corretto.\nScegli: non_esibito / non_idoneo / scaduto / non_dovuto")

    if "autorizzazione" in selected:
        _queue_control_question(state, "control_autorizzazione_status", "Licenza / autorizzazione NCC esibita: era regolare?\nScegli: regolare / non_esibita / non_autorizzato")
    else:
        _queue_control_question(state, "control_autorizzazione_status", "Licenza / autorizzazione NCC non esibita: scegli il caso corretto.\nScegli: non_esibita / non_autorizzato / regolare")

    if "foglio" in selected:
        _queue_control_question(state, "control_foglio_status", "Foglio di servizio esibito: scegli lo stato corretto.\nScegli: regolare / irregolare / assente / non_esibito")
    else:
        _queue_control_question(state, "control_foglio_status", "Foglio di servizio non esibito: scegli il caso corretto.\nScegli: non_esibito / assente / irregolare")

    _queue_control_question(state, "control_owner_type", "Intestatario/proprietario del mezzo: scegli il tipo.\nScegli: persona_fisica / cooperativa_srl / agenzia_viaggi / altro")
    _queue_control_question(state, "control_circulation_use", "Sul libretto / DU quale uso risulta?\nScegli: uso_terzi_ncc / uso_proprio / non_letto")
    _queue_control_question(state, "control_trip_nature", "Il servizio controllato sembra: NCC puro, pacchetto agenzia, navetta accessoria o dubbio?\nScegli: ncc_puro / agenzia_pacchetto / navetta_accessoria / dubbio")
    _queue_control_question(state, "control_rent_status", "Esito verifica RENT?\nScegli: si / no / non_verificato")
    _queue_control_question(state, "control_ruolo_status", "Esito verifica ruolo/albo conducenti?\nScegli: si / no / non_verificato")


def describe_control_violation(answers):
    parts = []
    if answers.get("patente_idonea") == "no":
        parts.append("il conducente guidava senza patente/titolo di guida idoneo")
    if answers.get("kb") == "no":
        parts.append("il conducente svolgeva il servizio senza CAP/KB/CQC richiesto")
    foglio = answers.get("foglio_status")
    if foglio == "assente":
        parts.append("il servizio NCC era svolto con foglio di servizio assente o non compilato")
    elif foglio == "irregolare":
        parts.append("il servizio NCC era svolto con foglio di servizio irregolare o incompleto")
    elif foglio == "non_esibito":
        parts.append("il foglio di servizio non veniva esibito all'atto del controllo")
    if answers.get("vehicle_authorized") == "no":
        parts.append("il veicolo risultava privo di licenza/autorizzazione NCC riferibile al servizio controllato")
    if answers.get("circulation_use") == "uso_proprio" and answers.get("trip_nature") == "ncc_puro":
        parts.append("il veicolo risultava in uso proprio ma veniva impiegato per trasporto di persone a pagamento")
    if answers.get("rent_registered") == "no":
        parts.append("il titolo/vettore non risultava iscritto al RENT")
    if answers.get("ruolo_conducenti") == "no":
        parts.append("il conducente non risultava iscritto al ruolo/albo conducenti")
    if not parts:
        return "Sono emersi elementi di possibile violazione in materia NCC da qualificare in base agli accertamenti svolti."
    text = "; ".join(parts)
    return text[0].upper() + text[1:] + "."


def build_recurrence_prompt(answers, key="recurrence"):

    descr = describe_control_violation(answers)
    if key == "recurrence_triennio":
        return (
            "VIOLAZIONE RISCONTRATA:\n"
            f"{descr}\n\n"
            "INQUADRAMENTO NORMATIVO:\n"
            "Ramo sanzionatorio art. 85 c.4 CdS (veicolo non adibito/autorizzato a NCC).\n\n"
            "Ora indica se si tratta di prima violazione o seconda nel triennio:\n"
            "first = prima\n"
            "second_3y = seconda nel triennio\n"
            "Rispondi con una di queste opzioni oppure usa i pulsanti qui sotto."
        )

    return (
        "VIOLAZIONE RISCONTRATA:\n"
        f"{descr}\n\n"
        "INQUADRAMENTO NORMATIVO:\n"
        "Ramo sanzionatorio art. 85 c.4-bis CdS, in relazione agli artt. 3 e/o 11 L. 21/1992.\n\n"
        "Ora indica la progressione della violazione nel quinquennio:\n"
        "first = prima\n"
        "2_5y = seconda nel quinquennio\n"
        "3_5y = terza nel quinquennio\n"
        "4plus_5y = quarta o successiva\n"
        "Rispondi con una di queste opzioni oppure usa i pulsanti qui sotto."
    )


def build_article_verification_prompt(answers, key, base_text):
    descr = describe_control_violation(answers)
    title = "VERIFICA MIRATA"
    framing = ""

    art85_keys = {"recurrence", "recurrence_triennio", "violation_type", "foglio_status", "booking", "public_waiting", "taxi_commune", "separate_payment"}
    kb_keys = {"kb", "control_kb_status"}
    patente_keys = {"patente_idonea", "control_patente_status"}
    auth_keys = {"vehicle_authorized", "control_autorizzazione_status"}
    rent_keys = {"rent_registered", "control_rent_status"}
    ruolo_keys = {"ruolo_conducenti", "control_ruolo_status"}
    uso_keys = {"circulation_use", "trip_nature", "owner_type", "control_circulation_use", "control_trip_nature", "control_owner_type"}
    incauto_keys = {"incauto_affidamento"}

    if key in art85_keys:
        title = "VIOLAZIONE IPOTIZZATA"
        if descr.startswith("Sono emersi elementi"):
            descr = "Possibile esercizio del servizio NCC in violazione delle modalità operative o delle prescrizioni autorizzative."
        framing = "INQUADRAMENTO PROVVISORIO:\nPossibile art. 85 c.4-bis CdS, in relazione agli artt. 3 e/o 11 L. 21/1992."
    elif key in kb_keys:
        descr = "Possibile irregolarità del titolo professionale richiesto (CAP/KB/CQC) per lo svolgimento del servizio."
        framing = "INQUADRAMENTO PROVVISORIO:\nVerifica necessaria per distinguere tra art. 180, art. 126 e art. 116 CdS."
    elif key in patente_keys:
        descr = "Possibile irregolarità della patente di guida rispetto al veicolo o al servizio controllato."
        framing = "INQUADRAMENTO PROVVISORIO:\nVerifica necessaria per distinguere tra art. 180, art. 126 e art. 116 CdS."
    elif key in auth_keys:
        descr = "Possibile irregolarità della licenza/autorizzazione NCC riferibile al servizio controllato."
        framing = "INQUADRAMENTO PROVVISORIO:\nVerifica necessaria per distinguere tra mera mancata esibizione documentale e violazione sostanziale in materia NCC."
    elif key in rent_keys:
        descr = "Occorre definire la posizione RENT del vettore/titolo controllato."
        framing = "INQUADRAMENTO PROVVISORIO:\nEsito utile per eventuali segnalazioni amministrative e per il quadro complessivo della regolarità NCC."
    elif key in ruolo_keys:
        descr = "Occorre definire la posizione del conducente rispetto al ruolo/albo conducenti."
        framing = "INQUADRAMENTO PROVVISORIO:\nEsito utile per eventuali segnalazioni all'ente competente e per la qualificazione finale del servizio."
    elif key in uso_keys:
        descr = "Occorre verificare la coerenza tra uso del veicolo risultante dal documento di circolazione e servizio effettivamente svolto."
        framing = "INQUADRAMENTO PROVVISORIO:\nPossibile art. 82 CdS oppure caso da approfondire se legato ad agenzia viaggi / trasporto accessorio."
    elif key in incauto_keys:
        descr = "Occorre verificare se ricorre l'incauto affidamento del veicolo a persona priva dei titoli richiesti."
        framing = "INQUADRAMENTO PROVVISORIO:\nPossibile art. 116 c.14 CdS in concorso con la violazione del conducente."
    else:
        framing = "INQUADRAMENTO PROVVISORIO:\nLa risposta serve a qualificare con precisione la norma applicabile."

    return (
        f"{title}:\n{descr}\n\n"
        f"{framing}\n\n"
        f"DOMANDA DI VERIFICA:\n{base_text}"
    )


def _apply_control_answer_to_state(state, key, value):
    answers = state.setdefault("answers", {})
    concurrent = state.setdefault("control_concurrent", [])
    notes = state.setdefault("control_notes", [])
    flags = state.setdefault("control_flags", {"segnalazioni": [], "verbale_additions": []})

    def add_note(n):
        _append_unique_local(notes, n)

    def add_flag(text, bucket="verbale_additions"):
        _append_unique_local(flags.setdefault(bucket, []), text)

    if key == "control_patente_status":
        if value == "valida":
            answers["patente_idonea"] = "si"
        elif value == "scaduta":
            answers["patente_idonea"] = "si"
            _append_unique_local(concurrent, "CDS_126_11")
            add_flag("Circolava alla guida del predetto veicolo con patente scaduta di validità; indicare la data di scadenza. La patente è ritirata e sarà inviata alla Prefettura-UTG competente.")
        elif value == "non_esibita":
            add_note("Patente non esibita all'atto del controllo: usare la voce documentale ex art. 180 se il titolo esiste ma non è stato mostrato.")
            _append_unique_local(concurrent, "180-01DOC")
        else:
            answers["patente_idonea"] = "no"

    elif key == "control_kb_status":
        if value in {"valido", "non_dovuto"}:
            answers["kb"] = "si"
        elif value == "scaduto":
            answers["kb"] = "si"
            _append_unique_local(concurrent, "CDS_126_11")
            add_flag("Circolava alla guida del predetto veicolo con CAP/KB/CQC scaduto di validità; indicare la data di scadenza. Il titolo è ritirato e sarà inviato all'UMC competente.")
        elif value == "non_esibito":
            add_note("KB / KA / CQC non esibito: se il titolo esiste ma non è stato mostrato, usare la voce documentale ex art. 180; se manca, ricorre l'art. 116 c.16 e c.18.")
            _append_unique_local(concurrent, "180-09")
        else:
            answers["kb"] = "no"

    elif key == "control_autorizzazione_status":
        if value == "regolare":
            answers["vehicle_authorized"] = "si"
        elif value == "non_esibita":
            answers["vehicle_authorized"] = "si"
            add_note("Licenza/autorizzazione NCC non esibita all'atto del controllo: usare la voce documentale ex art. 180 c.3 e c.7 secondo prontuario.")
            _append_unique_local(concurrent, "180-06")
        else:
            answers["vehicle_authorized"] = "no"
            answers["service_to_third"] = "si"

    elif key == "control_foglio_status":
        if value == "regolare":
            answers["foglio_status"] = "presente"
        elif value == "irregolare":
            answers["foglio_status"] = "irregolare"
            answers["violation_type"] = "art3_11"
        else:
            answers["foglio_status"] = value

    elif key == "control_owner_type":
        answers["owner_type"] = value
        if value == "agenzia_viaggi":
            add_note("Intestatario/organizzatore qualificato come agenzia viaggi: verificare se il trasporto è accessorio a pacchetto turistico documentato.")

    elif key == "control_circulation_use":
        answers["circulation_use"] = value
        if value == "uso_proprio":
            add_note("Dal libretto/DU risulta uso proprio: verificare compatibilità con il servizio effettivamente svolto e con l'eventuale titolo NCC.")

    elif key == "control_trip_nature":
        answers["trip_nature"] = value
        if value == "agenzia_pacchetto":
            add_note("Servizio dichiarato/emerso come inserito in pacchetto o escursione organizzata da agenzia: acquisire voucher, contratto, programma e prova del corrispettivo complessivo.")

    elif key in {"control_rent_status", "rent_registered"}:
        if value in {"si", "no", "non_verificato"}:
            answers["rent_registered"] = value

    elif key in {"control_ruolo_status", "ruolo_conducenti"}:
        if value in {"si", "no", "non_verificato"}:
            answers["ruolo_conducenti"] = value

    elif key == "recurrence":
        if value in {"first", "2_5y", "3_5y", "4plus_5y"}:
            answers["recurrence"] = value

    elif key == "recurrence_triennio":
        if value in {"first", "second_3y"}:
            answers["recurrence_triennio"] = value
            answers["recurrence"] = value

    elif key == "incauto_affidamento":
        if value in {"si", "no"}:
            answers["incauto_affidamento"] = value
save_user_states()

def _finalize_control(chat_id):
    state = user_states[chat_id]
    print(f"[DEBUG] finalize answers={state.get('answers', {})}")
    main_code, concurrent, notes, procedural_flags, ancillary_findings = decide_violation(state.get("answers", {}))

    for code in state.get("control_concurrent", []):
        _append_unique_local(concurrent, code)
    for note in state.get("control_notes", []):
        _append_unique_local(notes, note)
    for bucket, items in state.get("control_flags", {}).items():
        for item in items:
            _append_unique_local(procedural_flags.setdefault(bucket, []), item)

    if main_code:
        payload = build_final_payload(
            main_code,
            concurrent_codes=concurrent,
            extra_notes=notes,
            procedural_flags=procedural_flags,
            ancillary_findings=ancillary_findings
        )

        state["last_result_payload"] = payload
        state["last_result_main_code"] = main_code
        state["last_result_concurrent"] = concurrent
        state["last_result_flags"] = procedural_flags
        save_user_states()

        quick_text = payload["quick"]
        clear_case(chat_id)
        state = user_states.setdefault(chat_id, {})
        state["last_result_payload"] = payload
        state["last_result_main_code"] = main_code
        state["last_result_concurrent"] = concurrent
        state["last_result_flags"] = procedural_flags
        save_user_states()

        return quick_text

    if concurrent:
        result = format_partial_assessment(
            state.get("answers", {}),
            concurrent,
            notes,
            procedural_flags,
            ancillary_findings
        )
        clear_case(chat_id)
        return result

    result = "Dalla checklist documentale non emerge, allo stato, una violazione chiudibile automaticamente. Se il mezzo e il conducente sono in regola, il controllo può chiudersi senza contestazioni. In presenza di ulteriori elementi di fatto usa /caso oppure ripeti /controllo."
    clear_case(chat_id)
    return result

def control_additional_questions(answers):
    questions = []
    if answers.get("rent_registered") is None:
        questions.append({
            "key": "rent_registered",
            "text": "Il vettore/titolo risultava iscritto al RENT?\nRispondi: si / no"
        })
    if answers.get("ruolo_conducenti") is None:
        questions.append({
            "key": "ruolo_conducenti",
            "text": "Il conducente risultava iscritto al ruolo/albo conducenti quando richiesto?\nRispondi: si / no"
        })
    if answers.get("foglio_status") in {"assente", "irregolare"} and answers.get("recurrence") is None:
        questions.append({
            "key": "recurrence",
            "text": "Per il ramo art. 85 c.4-bis questa violazione è:\nfirst = prima\n2_5y = seconda nel quinquennio\n3_5y = terza nel quinquennio\n4plus_5y = quarta o successiva\nRispondi con una di queste opzioni."
        })
    if (answers.get("kb") == "no" or answers.get("patente_idonea") == "no") and answers.get("incauto_affidamento") is None:
        questions.append({
            "key": "incauto_affidamento",
            "text": "Il veicolo è stato affidato dal titolare o da altro responsabile a soggetto privo dei titoli richiesti?\nRispondi: si / no"
        })
    return questions


def next_control_question_or_result(chat_id):
    state = user_states[chat_id]
    queue = state.get("control_queue", [])
    if queue:
        q = queue.pop(0)
        state["mode"] = "control_followup"
        state["pending_question"] = q
        save_user_states()
        prompt = build_recurrence_prompt(state.get("answers", {}), q["key"]) if q["key"] in {"recurrence", "recurrence_triennio"} else build_article_verification_prompt(state.get("answers", {}), q["key"], q["text"])
        return prompt, q["key"]

    followup_questions = control_additional_questions(state.get("answers", {}))
    if followup_questions:
        q = followup_questions[0]
        state["mode"] = "control_followup"
        state["pending_question"] = q
        save_user_states()
        qa = state.setdefault("questions_asked", [])
        if q["key"] not in qa:
            qa.append(q["key"])
        save_user_states()    
        prompt = build_recurrence_prompt(state.get("answers", {}), q["key"]) if q["key"] in {"recurrence", "recurrence_triennio"} else build_article_verification_prompt(state.get("answers", {}), q["key"], q["text"])
        return prompt, q["key"]

    return _finalize_control(chat_id), None


def send_control_intro(chat_id):
    state = get_state(chat_id)
    return _control_text_from_state(state), build_control_docs_markup(state)


def begin_case_flow(chat_id):
    user_states[chat_id] = {
        "mode": "free_case",
        "free_text": "",
        "answers": {},
        "pending_question": None,
        "awaiting_external_consent": False,
        "awaiting_web_query": False,
        "questions_asked": [],
        "preset_name": None
    }
    save_user_states()


def begin_preset_case(chat_id, preset_name):
    begin_case_flow(chat_id)
    state = user_states[chat_id]
    state["preset_name"] = preset_name
    save_user_states()
    preset = PRESET_SCENARIOS[preset_name]
    first_response = process_case_description(chat_id, preset["text"])
    intro = (
        f"SCENARIO GUIDATO ATTIVATO: {preset['label']}\n\n"
        "Il bot ha già precaricato gli elementi tipici del caso e ora ti guida fino all'esito finale con sanzioni, diciture da verbale e comunicazioni conseguenti.\n"
    )
    return intro + "\n" + first_response

def process_case_description(chat_id, text):
    state = user_states[chat_id]
    state["free_text"] = text

    merge_detected_answers(state, text)
    save_user_states()

    case_key = match_case_from_text(text)
    main_code, concurrent, notes, procedural_flags, ancillary_findings = decide_violation(state["answers"])
    questions = missing_questions(state["answers"])

    if main_code and len(questions) == 0 and not should_offer_external_search(state["answers"], notes):
        level = confidence_level(state["answers"], main_code)
        result = format_multiple(main_code, concurrent, notes, level=level, procedural_flags=procedural_flags, ancillary_findings=ancillary_findings)
        clear_case(chat_id)
        return result

    if case_key and questions:
        q = questions[0]
        state["mode"] = "clarification"
        state["pending_question"] = q
        state["questions_asked"].append(q["key"])
        save_user_states()
        
        return (
            f"{format_case_hint(case_key)}\n\n"
            "Prima valutazione completata. Per chiudere correttamente il caso mi serve questo chiarimento:\n\n"
            f"{q['text']}"
        )

    if main_code and len(questions) <= 5 and len(questions) > 0:
        q = questions[0]
        state["mode"] = "clarification"
        state["pending_question"] = q
        state["questions_asked"].append(q["key"])
        save_user_states()
        return (
            "Esito preliminare probabile individuato con il database interno.\n"
            "Per completare il quadro sanzionatorio e le eventuali segnalazioni mi serve ancora questo dato:\n\n"
            f"{q['text']}"
        )

    if len(questions) == 0 and should_offer_external_search(state["answers"], notes):
        state["mode"] = "external_consent"
        state["awaiting_external_consent"] = True
        save_user_states()
        return ask_external_search_consent()

    if len(questions) == 0 and not main_code:
        result = format_partial_assessment(state["answers"], concurrent, notes, procedural_flags, ancillary_findings)
        clear_case(chat_id)
        return result

    if questions:
        q = questions[0]
        state["mode"] = "clarification"
        state["pending_question"] = q
        state["questions_asked"].append(q["key"])
        save_user_states()
        return (
            "Il bot non ha ancora dati sufficienti per una qualificazione affidabile.\n"
            "Ti faccio una domanda mirata per ricostruire tutte le possibili infrazioni:\n\n"
            f"{q['text']}"
        )

    state["mode"] = "external_consent"
    state["awaiting_external_consent"] = True
    save_user_states()
    return ask_external_search_consent()


def process_clarification(chat_id, text):
    state = user_states[chat_id]
    q = state.get("pending_question")

    if not q:
        clear_case(chat_id)
        return "Procedura annullata. Usa /caso per ricominciare."

    merge_detected_answers(state, text)
    value = parse_answer_for_key(q["key"], text)
    if value is None:
        return f"Risposta non valida.\n\n{q['text']}"

    target_key = q["key"]
    state["answers"][target_key] = value
    save_user_states()
    if q["key"] == "recurrence_triennio":
        state["answers"]["recurrence"] = value

    main_code, concurrent, notes, procedural_flags, ancillary_findings = decide_violation(state["answers"])
    questions = [item for item in missing_questions(state["answers"]) if item["key"] != q["key"]]

    if main_code and len(questions) == 0 and not should_offer_external_search(state["answers"], notes):
        level = confidence_level(state["answers"], main_code)
        result = format_multiple(main_code, concurrent, notes, level=level, procedural_flags=procedural_flags, ancillary_findings=ancillary_findings)
        clear_case(chat_id)
        return result

    if questions:
        next_q = questions[0]
        state["pending_question"] = next_q
        save_user_states()
        if next_q["key"] not in state["questions_asked"]:
            state["questions_asked"].append(next_q["key"])
        return f"Mi serve ancora questo chiarimento per avere il quadro completo:\n\n{next_q['text']}"

    if should_offer_external_search(state["answers"], notes):
        state["mode"] = "external_consent"
        state["awaiting_external_consent"] = True
        state["pending_question"] = None
        save_user_states()
        return ask_external_search_consent()

    if main_code:
        level = confidence_level(state["answers"], main_code)
        result = format_multiple(main_code, concurrent, notes, level=level, procedural_flags=procedural_flags, ancillary_findings=ancillary_findings)
        clear_case(chat_id)
        return result

    result = format_partial_assessment(state["answers"], concurrent, notes, procedural_flags, ancillary_findings)
    clear_case(chat_id)
    return result


def ask_external_search_consent():
    return (
        "Il database interno del bot non è sufficiente per chiudere il caso con affidabilità.\n\n"
        "Vuoi autorizzare una ricerca esterna su internet con supporto AI?\n"
        "Rispondi: si / no"
    )

def external_search_not_enabled_message():
    return (
        "Ricerca esterna autorizzata dall'operatore.\n\n"
        "ATTENZIONE:\n"
        "Il modulo AI/internet non è ancora collegato operativamente nel bot.\n"
        "Per ora occorre procedere con verifica manuale su fonti esterne.\n\n"
        "Quando il collegamento sarà attivo, il bot lo segnalerà espressamente."
    )

def manual_verification_message():
    return (
        "Ricerca esterna non autorizzata.\n\n"
        "Il caso non può essere chiuso con il solo database interno del bot.\n"
        "Procedere con verifica manuale su normativa vigente, prontuario e fonti esterne."
    )

def need_external_source_notice(answers):
    """Fallback prudenziale: segnala ricerca esterna solo nei casi davvero dubbi."""
    if not isinstance(answers, dict):
        return False
    if answers.get("owner_type") == "agenzia_viaggi" and answers.get("trip_nature") in {None, "dubbio", "agenzia_pacchetto"}:
        return True
    if answers.get("circulation_use") == "uso_proprio" and answers.get("service_to_third") in {None, "dubbio"}:
        return True
    return False

def should_offer_external_search(answers, notes):
    if need_external_source_notice(answers):
        return True

    if answers.get("service_to_third") == "dubbio":
        return True

    if answers.get("violation_type") == "none":
        return True

    if notes and len(notes) > 0:
        return True

    return False

def process_external_consent(chat_id, text):
    state = user_states[chat_id]
    t = text.strip().lower()

    if t not in {"si", "no"}:
        return "Rispondi solo con: si / no"

    if t == "no":
        clear_case(chat_id)
        return manual_verification_message()

    # t == "si"
    clear_case(chat_id)
    return external_search_not_enabled_message()

# =========================
# FLASK
# =========================

@app.route("/")
def home():
    return "NCC Sanzioni Bot attivo", 200

# =========================
# COMANDI ADMIN
# =========================

@bot.message_handler(commands=['approva'])
def approve_command(message):
    if not is_admin(message.from_user.id):
        return
    parts = message.text.strip().split()
    if len(parts) != 2 or not parts[1].isdigit():
        bot.reply_to(message, "Uso corretto: /approva ID")
        return
    user_id = int(parts[1])
    approve_user(user_id)
    bot.reply_to(message, f"Utente {user_id} autorizzato.")
    try:
        bot.send_message(user_id, "Accesso autorizzato dall'amministratore.\n\nOra puoi usare il bot.\nScrivi /start")
    except Exception:
        pass

@bot.message_handler(func=lambda m: (m.text or "").strip() == "Casi comuni porto")
def menu_port_common_cases_button(message):
    if not ensure_authorized(message):
        return

    bot.send_message(
        message.chat.id,
        "Seleziona una delle casistiche più frequenti in area porto.",
        reply_markup=build_port_common_cases_markup()
    )

@bot.message_handler(commands=['rifiuta'])
def reject_command(message):
    if not is_admin(message.from_user.id):
        return
    parts = message.text.strip().split()
    if len(parts) != 2 or not parts[1].isdigit():
        bot.reply_to(message, "Uso corretto: /rifiuta ID")
        return
    user_id = int(parts[1])
    reject_user(user_id)
    bot.reply_to(message, f"Utente {user_id} rifiutato.")
    try:
        bot.send_message(user_id, "La tua richiesta di accesso al bot non è stata approvata.")
    except Exception:
        pass

@bot.message_handler(commands=['pendenti'])
def pending_command(message):
    if not is_admin(message.from_user.id):
        return
    if not access_data["pending_users"]:
        bot.reply_to(message, "Nessuna richiesta pendente.")
        return
    lines = ["Richieste pendenti:"]
    for uid, data in access_data["pending_users"].items():
        username = f"@{data['username']}" if data.get("username") else "-"
        lines.append(f"- ID {uid} | {data.get('first_name', '-')} | {username}")
    bot.reply_to(message, "\n".join(lines))

@bot.message_handler(commands=['autorizzati'])
def authorized_command(message):
    if not is_admin(message.from_user.id):
        return
    bot.reply_to(message, "\n".join(format_authorized_users_lines(include_stats=False)))

@bot.message_handler(commands=['utenti'])
def utenti_command(message):
    if not is_admin(message.from_user.id):
        return
    bot.reply_to(message, "\n".join(format_authorized_users_lines(include_stats=True)))

@bot.message_handler(commands=['revoca'])
def revoke_command(message):
    if not is_admin(message.from_user.id):
        return
    parts = message.text.strip().split()
    if len(parts) != 2 or not parts[1].isdigit():
        bot.reply_to(message, "Uso corretto: /revoca ID")
        return
    user_id = int(parts[1])
    ok = revoke_user(user_id)
    if not ok:
        bot.reply_to(message, "Non puoi revocare l'accesso all'admin.")
        return
    bot.reply_to(message, f"Accesso revocato all'utente {user_id}.")
    try:
        bot.send_message(user_id, "Il tuo accesso al bot è stato revocato.")
    except Exception:
        pass

@bot.message_handler(commands=['restartbot'])
def restartbot_command(message):
    if not is_admin(message.from_user.id):
        return

    ok, msg = restart_render_service()
    if ok:
        bot.reply_to(
            message,
            "Richiesta di riavvio inviata a Render.\n\n"
            "Attendi circa 30-60 secondi per il completo ripristino del servizio.\n\n"
            f"Dettaglio: {msg}"
        )
    else:
        bot.reply_to(message, f"Riavvio non riuscito.\n\n{msg}")

@bot.message_handler(commands=['deploybot'])
def deploybot_command(message):
    if not is_admin(message.from_user.id):
        return

    ok, msg = deploy_render_service()
    if ok:
        bot.reply_to(
            message,
            "Richiesta di deploy inviata a Render.\n\n"
            "Attendi circa 1-2 minuti per il completamento.\n\n"
            f"Dettaglio: {msg}"
        )
    else:
        bot.reply_to(message, f"Deploy non riuscito.\n\n{msg}")

def confidence_level(answers, main_code):
    score = 0

    if answers.get("vehicle_authorized") is not None:
        score += 1
    if answers.get("service_to_third") is not None:
        score += 1
    if answers.get("kb") is not None:
        score += 1
    if answers.get("booking") is not None:
        score += 1
    if answers.get("separate_payment") is not None:
        score += 1
    if main_code is not None:
        score += 2

    if score >= 6:
        return "molto probabile"
    if score >= 4:
        return "probabile"
    return "da approfondire"

# =========================
# COMANDI BOT
# =========================

@bot.message_handler(commands=['start'])
def start_command(message):
    uid = message.from_user.id
    send_welcome_media(message.chat.id)

    if not is_admin(uid) and not is_authorized(uid):
        if not is_pending(uid):
            add_pending(message.from_user)
            notify_admin_new_request(message.from_user)
        bot.reply_to(message, request_access_text())
        return

    track_authorized_usage(message.from_user, "/start")

    bot.send_message(
        message.chat.id,
        authorized_start_text(uid),
        reply_markup=build_main_menu()
    )

@bot.message_handler(commands=['help'])
def help_command(message):
    if not ensure_authorized(message):
        return

    text = (
        "Funzioni disponibili:\n\n"
        "• Inserisci un caso NCC = analisi libera del fatto\n"
        "• Checklist documentale = controllo guidato con pulsanti\n"
        "• Controlli operativi = verifiche pratiche sul posto\n"
        "• Documenti da controllare = elenco documenti da richiedere o verificare\n"
        "• Norme principali = riferimenti normativi NCC con pulsanti per gli articoli\n"
        "• Verifica targa = ricerca targa nell'archivio NCC\n"
        "• Reset = annulla la procedura in corso\n\n"
        "Puoi usare i pulsanti sotto la chat oppure il menu comandi di Telegram."
    )

    if is_admin(message.from_user.id):
        text += (
            "\n\nComandi admin:\n"
            "/restartbot = riavvia il servizio su Render\n"
            "/deploybot = avvia un deploy su Render"
        )

    bot.send_message(
        message.chat.id,
        text,
        reply_markup=build_main_menu()
    )

@bot.message_handler(commands=['norme'])
def norme_command(message):
    if not ensure_authorized(message):
        return

    text = format_norme_from_db() + "\n\nSeleziona un articolo per leggerlo:"
    
    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.row(
        types.InlineKeyboardButton("Art. 85 CdS", callback_data="article:art85"),
        types.InlineKeyboardButton("Art. 116 CdS", callback_data="article:art116"),
    )
    markup.row(
        types.InlineKeyboardButton("Art. 3 L.21/1992", callback_data="article:art3l21"),
        types.InlineKeyboardButton("Art. 11 L.21/1992", callback_data="article:art11l21"),
    )
    markup.row(
        types.InlineKeyboardButton("Art. 180 CdS", callback_data="article:art180"),
        types.InlineKeyboardButton("Art. 126 CdS", callback_data="article:art126"),
    )

    bot.reply_to(message, text, reply_markup=markup)

@bot.message_handler(commands=['targa'])
def targa_command(message):
    if not ensure_authorized(message):
        return

    text = (message.text or "").strip()
    parts = text.split(maxsplit=1)

    if len(parts) > 1 and parts[1].strip():
        plate = parts[1].strip()
        result = lookup_plate_in_registry(plate)

        if result.get("ok") and not result.get("found"):
            bot.reply_to(
                message,
                result.get("message", "Targa non censita.") + "\n\n"
                "Se vuoi, premi il pulsante qui sotto per inviare la richiesta di censimento.",
                reply_markup=build_plate_not_found_markup(result.get("plate", plate))
            )
            return

        bot.reply_to(message, result.get("message", "Errore nella ricerca targa."))
        return

    begin_plate_lookup_flow(message.chat.id)
    bot.reply_to(
        message,
        "Inserisci la targa del mezzo da verificare.\n\n"
        "Esempio: AB123CD\n"
        "Il bot controllerà l'archivio Excel aggiornato nel repository e ti dirà se il mezzo è adibito o meno al servizio NCC."
    )


@bot.message_handler(commands=['documenti'])
def documenti_command(message):
    if not ensure_authorized(message):
        return
    bot.reply_to(message, format_documenti_from_db())

@bot.message_handler(commands=['art85'])
def art85_command(message):
    if not ensure_authorized(message):
        return

    text = format_articolo("art85")
    state = get_state(message.chat.id)
    if state and state.get("mode") in ["clarification", "external_consent", "control_followup"]:
        text += "\n\nLa procedura guidata è ancora attiva. Dopo la lettura dell'articolo, riprendi rispondendo alla domanda precedente oppure usa /reset."

    bot.reply_to(message, text)

@bot.message_handler(commands=['art116'])
def art116_command(message):
    if not ensure_authorized(message):
        return

    text = format_articolo("art116")
    state = get_state(message.chat.id)
    if state and state.get("mode") in ["clarification", "external_consent", "control_followup"]:
        text += "\n\nLa procedura guidata è ancora attiva. Dopo la lettura dell'articolo, riprendi rispondendo alla domanda precedente oppure usa /reset."

    bot.reply_to(message, text)

@bot.message_handler(commands=['art3l21'])
def art3l21_command(message):
    if not ensure_authorized(message):
        return

    text = format_articolo("art3l21")
    state = get_state(message.chat.id)
    if state and state.get("mode") in ["clarification", "external_consent", "control_followup"]:
        text += "\n\nLa procedura guidata è ancora attiva. Dopo la lettura dell'articolo, riprendi rispondendo alla domanda precedente oppure usa /reset."

    bot.reply_to(message, text)

@bot.message_handler(commands=['art11l21'])
def art11l21_command(message):
    if not ensure_authorized(message):
        return

    text = format_articolo("art11l21")
    state = get_state(message.chat.id)
    if state and state.get("mode") in ["clarification", "external_consent", "control_followup"]:
        text += "\n\nLa procedura guidata è ancora attiva. Dopo la lettura dell'articolo, riprendi rispondendo alla domanda precedente oppure usa /reset."

    bot.reply_to(message, text)

@bot.message_handler(commands=['art180'])
def art180_command(message):
    if not ensure_authorized(message):
        return

    text = format_articolo("art180")
    state = get_state(message.chat.id)
    if state and state.get("mode") in ["clarification", "external_consent", "control_followup"]:
        text += "\n\nLa procedura guidata è ancora attiva. Dopo la lettura dell'articolo, riprendi rispondendo alla domanda precedente oppure usa /reset."

    bot.reply_to(message, text, reply_markup=build_violation_markup_for_article("art180"))


@bot.message_handler(commands=['art126'])
def art126_command(message):
    if not ensure_authorized(message):
        return

    text = format_articolo("art126")
    state = get_state(message.chat.id)
    if state and state.get("mode") in ["clarification", "external_consent", "control_followup"]:
        text += "\n\nLa procedura guidata è ancora attiva. Dopo la lettura dell'articolo, riprendi rispondendo alla domanda precedente oppure usa /reset."

    bot.reply_to(message, text)


@bot.message_handler(commands=['verbale'])
def verbale_command(message):
    if not ensure_authorized(message):
        return
    bot.reply_to(message, format_verbale_template())

@bot.message_handler(commands=['checklist'])
def checklist_command(message):
    if not ensure_authorized(message):
        return
    bot.reply_to(message, format_checklist_from_db())

@bot.message_handler(commands=['controllo'])
def controllo_command(message):
    if not ensure_authorized(message):
        return
    begin_control_flow(message.chat.id)
    text, markup = send_control_intro(message.chat.id)
    bot.send_message(message.chat.id, text, reply_markup=markup)

@bot.message_handler(commands=['reset'])
def reset_command(message):
    if not ensure_authorized(message):
        return
    clear_case(message.chat.id)
    bot.reply_to(message, "Procedura annullata.")

@bot.message_handler(commands=['riattiva'])
def riattiva_command(message):
    # questo comando deve essere disponibile anche a utenti autorizzati normali
    if not ensure_authorized(message):
        return

    bot.reply_to(message, public_wake_up_message(), disable_web_page_preview=True)

@bot.message_handler(commands=['caso'])
def caso_command(message):
    if not ensure_authorized(message):
        return
    begin_case_flow(message.chat.id)
    bot.reply_to(
        message,
        "Descrivi liberamente la situazione in un solo messaggio. Per i controlli documentali standard è preferibile /controllo, che usa pulsanti e riduce le domande inutili.\nIl bot cercherà tutte le possibili infrazioni; se mancano dati ti farà domande mirate fino a chiudere il quadro.\n\n"
        "Esempio:\n"
        "veicolo ncc fermo al porto, foglio di servizio assente, rent non registrato, conducente senza kb, mezzo affidato dal titolare"
    )

@bot.message_handler(commands=['porto'])
def porto_command(message):
    if not ensure_authorized(message):
        return
    bot.reply_to(message, begin_preset_case(message.chat.id, "porto"))

@bot.message_handler(commands=['aeroporto'])
def aeroporto_command(message):
    if not ensure_authorized(message):
        return
    bot.reply_to(message, begin_preset_case(message.chat.id, "aeroporto"))

@bot.message_handler(commands=['stazione'])
def stazione_command(message):
    if not ensure_authorized(message):
        return
    bot.reply_to(message, begin_preset_case(message.chat.id, "stazione"))

@bot.message_handler(commands=['hotel'])
def hotel_command(message):
    if not ensure_authorized(message):
        return
    bot.reply_to(message, begin_preset_case(message.chat.id, "hotel"))

@bot.message_handler(commands=['navetta'])
def navetta_command(message):
    if not ensure_authorized(message):
        return
    bot.reply_to(message, begin_preset_case(message.chat.id, "navetta"))

@bot.message_handler(commands=['casiporto'])
def casi_porto_command(message):
    if not ensure_authorized(message):
        return

    bot.send_message(
        message.chat.id,
        "Seleziona una delle casistiche più frequenti in area porto.",
        reply_markup=build_port_common_cases_markup()
    )

@bot.message_handler(func=lambda m: (m.text or "").strip() == "Inserisci un caso NCC")
def menu_caso_button(message):
    caso_command(message)

@bot.message_handler(func=lambda m: (m.text or "").strip() == "Checklist documentale")
def menu_controllo_button(message):
    controllo_command(message)

@bot.message_handler(func=lambda m: (m.text or "").strip() == "Controlli operativi")
def menu_checklist_button(message):
    checklist_command(message)

@bot.message_handler(func=lambda m: (m.text or "").strip() == "Documenti da controllare")
def menu_documenti_button(message):
    documenti_command(message)

@bot.message_handler(func=lambda m: (m.text or "").strip() == "Norme principali")
def menu_norme_button(message):
    norme_command(message)

@bot.message_handler(func=lambda m: (m.text or "").strip() == "Verifica targa")
def menu_targa_button(message):
    if not ensure_authorized(message):
        return

    begin_plate_lookup_flow(message.chat.id)
    bot.reply_to(
        message,
        "Inserisci la targa del mezzo da verificare.\n\n"
        "Esempio: AB123CD\n"
        "Il bot controllerà l'archivio Excel aggiornato nel repository e ti dirà se il mezzo è adibito o meno al servizio NCC."
    )

@bot.message_handler(func=lambda m: (m.text or "").strip() == "Reset")
def menu_reset_button(message):
    reset_command(message)

@bot.message_handler(func=lambda m: (m.text or "").strip() == "Help")
def menu_help_button(message):
    help_command(message)

@bot.callback_query_handler(func=lambda call: str(call.data).startswith("ctrl_doc_toggle:"))
def control_doc_toggle_callback(call):
    chat_id = call.message.chat.id
    state = get_state(chat_id)
    if not state or state.get("mode") != "control_docs":
        try:
            bot.answer_callback_query(call.id, "Nessun controllo documentale attivo")
        except Exception:
            pass
        return

    doc_id = str(call.data).split(":", 1)[1].strip()
    selected = state.setdefault("selected_docs", [])
    if doc_id in selected:
        selected.remove(doc_id)
    else:
        selected.append(doc_id)
    save_user_states()
    try:
        bot.answer_callback_query(call.id, "Selezione aggiornata")
    except Exception:
        pass
    bot.edit_message_text(_control_text_from_state(state), chat_id, call.message.message_id, reply_markup=build_control_docs_markup(state))


@bot.callback_query_handler(func=lambda call: str(call.data).startswith("porto_case:"))
def porto_case_callback(call):
    chat_id = call.message.chat.id
    key = str(call.data).split(":", 1)[1].strip()

    try:
        bot.answer_callback_query(call.id, "Caso porto selezionato")
    except Exception:
        pass

    if key == "altro":
        text = begin_preset_case(chat_id, "porto")
        bot.send_message(chat_id, text)
        return

    started = begin_port_common_case(chat_id, key)
    if not started:
        bot.send_message(chat_id, "Caso porto non riconosciuto.")
        return

    prompt, question_key = started
    markup = build_combined_markup(question_key=question_key, force_ctrl_answer=False)
    bot.send_message(chat_id, prompt, reply_markup=markup)

@bot.callback_query_handler(func=lambda call: (call.data or "").startswith("plate_report:"))

def plate_report_callback(call):
    try:
        plate = (call.data or "").split(":", 1)[1].strip()

        notify_admin_missing_plate(call.from_user, plate)

        bot.answer_callback_query(
            call.id,
            "Richiesta inviata all'amministratore."
        )

        bot.send_message(
            call.message.chat.id,
            f"Richiesta inviata per la targa {plate}.\n\n"
            "La targa verrà verificata e, se confermata, censita nell'archivio."
        )
    except Exception as e:
        bot.answer_callback_query(call.id, "Errore nell'invio della richiesta.")
        try:
            bot.send_message(
                call.message.chat.id,
                f"Errore nella segnalazione: {e}"
            )
        except Exception:
            pass

@bot.callback_query_handler(func=lambda call: str(call.data) == "ctrl_doc_cancel")
def control_doc_cancel_callback(call):
    clear_case(call.message.chat.id)
    try:
        bot.answer_callback_query(call.id, "Controllo annullato")
    except Exception:
        pass
    bot.send_message(call.message.chat.id, "Procedura /controllo annullata.")


@bot.callback_query_handler(func=lambda call: str(call.data) == "ctrl_doc_done")
def control_doc_done_callback(call):
    chat_id = call.message.chat.id
    state = get_state(chat_id)

    if not state or state.get("mode") != "control_docs":
        try:
            bot.answer_callback_query(call.id, "Nessun controllo documentale attivo")
        except Exception:
            pass
        return

    apply_control_defaults_from_selection(state)
    build_control_queue(state)
    save_user_states()

    try:
        bot.answer_callback_query(call.id, "Documenti acquisiti")
    except Exception:
        pass

    try:
        print(f"[DEBUG] ctrl_doc_done answers={state.get('answers', {})}")
        result, qkey = next_control_question_or_result(chat_id)
        print(f"[DEBUG] ctrl_doc_done next qkey={qkey}")
        print(f"[DEBUG] ctrl_doc_done result_len={len(result) if result else 0}")

        if qkey:
            markup = build_combined_markup([], qkey, force_ctrl_answer=True)
        else:
            state_after = get_state(chat_id)
            payload = state_after.get("last_result_payload") if state_after else None
            main_code = state_after.get("last_result_main_code") if state_after else None
            concurrent_codes = state_after.get("last_result_concurrent", []) if state_after else []
            flags = state_after.get("last_result_flags", {}) if state_after else {}

            markup = build_final_result_markup(payload)
            if not markup:
                markup = build_pdf_markup(main_code, concurrent_codes, flags) or build_article_markup(
                    infer_article_keys_from_text(result)
                )

        send_long_message(
            chat_id,
            result,
            reply_markup=markup,
            disable_web_page_preview=True
        )
        
    except Exception as e:
        print(f"ERRORE control_doc_done_callback: {e}")
        print(traceback.format_exc())
        bot.send_message(chat_id, f"Errore interno nel controllo documentale: {e}")

@bot.callback_query_handler(func=lambda call: str(call.data).startswith("ctrl_answer:"))
def control_answer_callback(call):
    chat_id = call.message.chat.id
    state = get_state(chat_id)

    if not state or state.get("mode") != "control_followup":
        try:
            bot.answer_callback_query(call.id, "Nessuna domanda attiva")
        except Exception:
            pass
        return

    value = str(call.data).split(":", 1)[1].strip()
    q = state.get("pending_question")

    if not q:
        clear_case(chat_id)
        bot.send_message(chat_id, "Procedura annullata. Usa /controllo per ricominciare.")
        return

    try:
        print(f"[DEBUG] ctrl_answer key={q['key']} value={value}")
        print(f"[DEBUG] answers_before={state.get('answers', {})}")
        _apply_control_answer_to_state(state, q["key"], value)
        state["pending_question"] = None
        save_user_states()
        print(f"[DEBUG] answers_after={state.get('answers', {})}")

        try:
            bot.answer_callback_query(call.id, "Risposta acquisita")
        except Exception:
            pass

        result, qkey = next_control_question_or_result(chat_id)
        print(f"[DEBUG] next_control qkey={qkey}")
        print(f"[DEBUG] result_len={len(result) if result else 0}")

        if qkey:
            markup = build_combined_markup([], qkey, force_ctrl_answer=True)
        else:
            state_after = get_state(chat_id)
            payload = state_after.get("last_result_payload") if state_after else None
            main_code = state_after.get("last_result_main_code") if state_after else None
            concurrent_codes = state_after.get("last_result_concurrent", []) if state_after else []
            flags = state_after.get("last_result_flags", {}) if state_after else {}

            markup = build_final_result_markup(payload)
            if not markup:
                markup = build_pdf_markup(main_code, concurrent_codes, flags) or build_article_markup(
                    infer_article_keys_from_text(result)
                )

        send_long_message(
            chat_id,
            result,
            reply_markup=markup,
            disable_web_page_preview=True
        )

    except Exception as e:
        print(f"ERRORE control_answer_callback: {e}")
        print(traceback.format_exc())
        try:
            bot.answer_callback_query(call.id, "Errore nel flusso")
        except Exception:
            pass
        bot.send_message(chat_id, f"Errore interno nel flusso risposte: {e}")

@bot.callback_query_handler(func=lambda call: str(call.data).startswith("final:"))
def final_result_callback(call):
    chat_id = call.message.chat.id
    state = get_state(chat_id)

    if not state:
        try:
            bot.answer_callback_query(call.id, "Nessun esito finale disponibile")
        except Exception:
            pass
        return

    payload = state.get("last_result_payload")
    if not payload:
        try:
            bot.answer_callback_query(call.id, "Nessun esito finale disponibile")
        except Exception:
            pass
        return

    action = str(call.data).split(":", 1)[1].strip()

    text = None
    if action == "quick":
        text = payload.get("quick")
    elif action == "accessori":
        text = payload.get("accessori")
    elif action == "comunicazioni":
        text = payload.get("comunicazioni")
    elif action == "articoli":
        text = payload.get("articoli")
    elif action == "v1":
        verbali = payload.get("verbali", [])
        text = verbali[0] if len(verbali) >= 1 else "Verbale 1 non disponibile."
    elif action == "v2":
        verbali = payload.get("verbali", [])
        text = verbali[1] if len(verbali) >= 2 else "Verbale 2 non disponibile."
    elif action == "v3":
        verbali = payload.get("verbali", [])
        text = verbali[2] if len(verbali) >= 3 else "Verbale 3 non disponibile."
    elif action == "v4":
        verbali = payload.get("verbali", [])
        text = verbali[3] if len(verbali) >= 4 else "Verbale 4 non disponibile."
    else:
        text = "Sezione non disponibile."

    try:
        bot.answer_callback_query(call.id, "Sezione aperta")
    except Exception:
        pass

    send_long_message(
        chat_id,
        text,
        reply_markup=build_final_result_markup(payload),
        disable_web_page_preview=True
    )

@bot.callback_query_handler(func=lambda call: str(call.data).startswith("article:"))
def article_callback(call):
    key = str(call.data).split(":", 1)[1].strip()
    text = format_articolo(key)
    markup = build_violation_markup_for_article(key)

    try:
        bot.answer_callback_query(call.id, "Articolo richiamato")
    except Exception:
        pass

    send_long_message(
        call.message.chat.id,
        text,
        reply_markup=markup,
        disable_web_page_preview=True
    )


@bot.callback_query_handler(func=lambda call: str(call.data).startswith("viol:"))
def violation_callback(call):
    code = normalize_violation_code(str(call.data).split(":", 1)[1].strip())
    if not code:
        try:
            bot.answer_callback_query(call.id, "Voce non disponibile")
        except Exception:
            pass
        return

    try:
        bot.answer_callback_query(call.id, "Voce sanzionatoria richiamata")
    except Exception:
        pass

    send_long_message(
        call.message.chat.id,
        format_compact_violation(code),
        disable_web_page_preview=True
    )


@bot.callback_query_handler(func=lambda call: str(call.data).startswith("answer:"))
def answer_callback(call):
    chat_id = call.message.chat.id
    state = get_state(chat_id)

    if not state:
        try:
            bot.answer_callback_query(call.id, "Nessuna domanda attiva")
        except Exception:
            pass
        return

    value = str(call.data).split(":", 1)[1].strip()

    try:
        if state.get("mode") == "clarification":
            response = process_clarification(chat_id, value)
            state = get_state(chat_id)
            question_key = None
            article_keys = infer_article_keys_from_text(response)

            if state and state.get("mode") == "clarification" and state.get("pending_question"):
                question_key = state["pending_question"].get("key")

            markup = build_combined_markup(article_keys, question_key=question_key)

            try:
                bot.answer_callback_query(call.id, "Risposta acquisita")
            except Exception:
                pass

            send_long_message(
                chat_id,
                response,
                reply_markup=markup,
                disable_web_page_preview=True
            )
            return

        if state.get("mode") == "porto_common_followup":
            result, payload, question_key = process_port_common_followup(chat_id, value)

            try:
                bot.answer_callback_query(call.id, "Risposta acquisita")
            except Exception:
                pass

            if payload:
                markup = build_final_result_markup(payload)
                if not markup:
                    state_after = get_state(chat_id) or {}
                    main_code = state_after.get("last_result_main_code")
                    concurrent_codes = state_after.get("last_result_concurrent", [])
                    flags = state_after.get("last_result_flags", {})
                    markup = build_pdf_markup(main_code, concurrent_codes, flags) or build_article_markup(
                        get_article_keys_for_result(main_code, concurrent_codes)
                    )

                send_long_message(
                    chat_id,
                    result,
                    reply_markup=markup,
                    disable_web_page_preview=True
                )
                return

            markup = build_combined_markup(question_key=question_key) if question_key else None

            send_long_message(
                chat_id,
                result,
                reply_markup=markup,
                disable_web_page_preview=True
            )
            return

        try:
            bot.answer_callback_query(call.id, "Nessuna domanda attiva")
        except Exception:
            pass

    except Exception as e:
        try:
            bot.answer_callback_query(call.id, "Errore nel flusso")
        except Exception:
            pass
        bot.send_message(chat_id, f"Errore interno nel flusso risposte: {e}")

# =========================
# MESSAGGI GENERICI
# =========================

@bot.message_handler(func=lambda m: True)
def all_messages(message):
    text = (message.text or "").strip()

    # I comandi /... devono essere gestiti solo dai rispettivi handler dedicati,
    # senza interrompere il flusso del caso guidato o del controllo documentale.
    if text.startswith("/"):
        return

    if not ensure_authorized(message):
        return

    chat_id = message.chat.id
    state = get_state(chat_id)

    if not state:
        bot.reply_to(message, "Usa /controllo per la checklist documentale guidata, /caso per il testo libero, /targa per verificare una targa, oppure uno scenario guidato tra /porto /aeroporto /stazione /hotel /navetta.")
        return

    mode = state.get("mode")

    if mode == "free_case":
        response = process_case_description(chat_id, text)
        reply_with_article_buttons(message, response)
        return

    if mode == "clarification":
        response = process_clarification(chat_id, text)
        reply_with_article_buttons(message, response)
        return

    if mode == "plate_lookup":
        result = process_plate_lookup(chat_id, text)

        if result.get("ok") and not result.get("found"):
            bot.reply_to(
                message,
                result.get("message", "Targa non censita.") + "\n\n"
                "Se vuoi, premi il pulsante qui sotto per inviare la richiesta di censimento.",
                reply_markup=build_plate_not_found_markup(result.get("plate", text))
            )
        else:
            bot.reply_to(message, result.get("message", "Errore nella ricerca targa."))
        return

    if mode == "porto_common_followup":
        result, payload, question_key = process_port_common_followup(chat_id, text)

        if payload:
            markup = build_final_result_markup(payload)
            if not markup:
                state_after = get_state(chat_id) or {}
                main_code = state_after.get("last_result_main_code")
                concurrent_codes = state_after.get("last_result_concurrent", [])
                flags = state_after.get("last_result_flags", {})
                markup = build_pdf_markup(main_code, concurrent_codes, flags) or build_article_markup(
                    get_article_keys_for_result(main_code, concurrent_codes)
                )
            send_long_message(chat_id, result, reply_markup=markup, disable_web_page_preview=True)
            return

        markup = build_combined_markup(question_key=question_key) if question_key else None
        bot.reply_to(message, result, reply_markup=markup)
        return

    if mode == "control_docs":
        bot.reply_to(message, "Usa i pulsanti del controllo documentale oppure /reset per annullare.")
        return

    if mode == "control_followup":
        q = state.get("pending_question")
        if not q:
            clear_case(chat_id)
            bot.reply_to(message, "Procedura annullata. Usa /controllo per ricominciare.")
            return
        value = parse_answer_for_key(q["key"], text)
        if value is None:
            markup = build_combined_markup([], q["key"])
            bot.reply_to(message, f"Risposta non valida.\n\n{q['text']}", reply_markup=markup)
            return
        _apply_control_answer_to_state(state, q["key"], value)
        state["pending_question"] = None
        save_user_states()
        result, qkey = next_control_question_or_result(chat_id)

        if qkey:
            markup = build_combined_markup([], qkey, force_ctrl_answer=True)
        else:
            state_after = get_state(chat_id)
            payload = state_after.get("last_result_payload") if state_after else None
            main_code = state_after.get("last_result_main_code") if state_after else None
            concurrent_codes = state_after.get("last_result_concurrent", []) if state_after else []
            flags = state_after.get("last_result_flags", {}) if state_after else {}

            markup = build_final_result_markup(payload)
            if not markup:
                markup = build_pdf_markup(main_code, concurrent_codes, flags) or build_article_markup(
                    infer_article_keys_from_text(result)
                )

        send_long_message(chat_id, result, reply_markup=markup, disable_web_page_preview=True)
        return

    if mode == "external_consent":
        response = process_external_consent(chat_id, text)
        bot.reply_to(message, response)
        return

    bot.reply_to(message, "Usa /caso per descrivere il fatto oppure /reset per annullare la procedura.")

# =========================
# AVVIO BOT
# =========================


def setup_bot_commands():
    commands = [
        types.BotCommand("start", "Avvia il bot"),
        types.BotCommand("reset", "Annulla procedura"),
        types.BotCommand("riattiva", "Riattiva il servizio"),
        types.BotCommand("help", "Come usare il bot"),
    ]
    try:
        bot.set_my_commands(commands)
    except Exception as e:
        print(f"Errore setup comandi bot: {e}")
        
def run_bot():
    print("AVVIO BOT TELEGRAM...")
    if not TOKEN:
        raise RuntimeError("TOKEN mancante nelle variabili environment di Render")

    while True:
        try:
            bot.remove_webhook()
            setup_bot_commands()
            print("Polling Telegram in avvio...")
            bot.infinity_polling(
                timeout=30,
                long_polling_timeout=30,
                skip_pending=True
            )
        except Exception as e:
            import time
            import traceback
            print(f"ERRORE AVVIO/POLLING BOT: {e}")
            print(traceback.format_exc())
            time.sleep(5)

if __name__ == "__main__":
    threading.Thread(target=run_bot, daemon=True).start()
    port = int(os.getenv("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
